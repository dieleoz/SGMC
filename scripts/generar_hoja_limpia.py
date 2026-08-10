# -*- coding: utf-8 -*-
"""Genera BD/Modelo_Datos_LIMPIO.xlsx: la hoja tal como la declara el modelo.

Por que existe
--------------
El modelo se definio y despues se heredo la hoja vieja tal cual, con 49 columnas
que el modelo no usa. En vez de generar la hoja desde el modelo, se escribio
documentacion para gestionar la basura: anexos de ocultacion, listas de trampas,
COLUMNAS_SIN_DECIDIR.

Todo eso desaparece si la hoja se genera del modelo.

Que hace
--------
Para cada tabla de MODELO crea una pestana con EXACTAMENTE sus columnas, en su
orden, y migra los datos del libro origen emparejando por nombre de columna. Lo
que el modelo no declara, no viaja.

Uso:  python scripts/generar_hoja_limpia.py "BD/Modelo_Datos_09082026.xlsx"
"""
import os
import sys

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "scripts"))
from modelo_objetivo import MODELO

try:
    import openpyxl
except ImportError:
    print("Falta openpyxl."); sys.exit(2)

origen = sys.argv[1] if len(sys.argv) > 1 else "BD/Modelo_Datos_09082026.xlsx"
if not os.path.isabs(origen):
    origen = os.path.join(RAIZ, origen)
salida = os.path.join(RAIZ, "BD", "Modelo_Datos_LIMPIO.xlsx")

src = openpyxl.load_workbook(origen, data_only=True, read_only=True)
dst = openpyxl.Workbook()
dst.remove(dst.active)

resumen = []
perdidas = []

for tabla in MODELO:
    cols = [c["nombre"] for c in MODELO[tabla]["columnas"]]
    ws = dst.create_sheet(tabla)
    ws.append(cols)

    if tabla not in src.sheetnames:
        resumen.append((tabla, len(cols), 0, "pestana nueva, sin datos de origen"))
        continue

    o = src[tabla]
    cab = [str(c.value).strip() if c.value is not None else "" for c in next(o.iter_rows(min_row=1, max_row=1))]
    idx = {n: i for i, n in enumerate(cab)}

    # columnas del origen que NO viajan
    fuera = [n for n in cab if n and n not in cols]
    if fuera:
        perdidas.append((tabla, fuera))

    n = 0
    for r in o.iter_rows(min_row=2, values_only=True):
        if not r or all(v in (None, "") for v in r):
            continue
        fila = []
        for c in cols:
            i = idx.get(c)
            fila.append(r[i] if i is not None and i < len(r) else None)
        if all(v in (None, "") for v in fila):
            continue
        ws.append(fila)
        n += 1
    resumen.append((tabla, len(cols), n, ""))

dst.save(salida)

print("Generado:", salida)
print()
print("%-26s %8s %8s" % ("TABLA", "COLS", "FILAS"))
for t, c, f, nota in sorted(resumen):
    print("%-26s %8d %8d  %s" % (t, c, f, nota))
print()
print("Pestanas: %d   ·   Columnas: %d   ·   Filas: %d"
      % (len(resumen), sum(x[1] for x in resumen), sum(x[2] for x in resumen)))
print()
if perdidas:
    total = sum(len(f) for _, f in perdidas)
    print("=== %d columnas del origen que NO viajan (el modelo no las declara) ===" % total)
    for t, fuera in sorted(perdidas):
        print("  %-24s %s" % (t, " · ".join(sorted(fuera))))
