# -*- coding: utf-8 -*-
"""Comprueba que los DATOS de la hoja sostienen lo que el modelo declara.

El sexto verificador, y nace del unico hueco que los otros cinco compartian:
NINGUNO ABRE EL ARCHIVO DE DATOS PARA MIRAR SI LAS COLUMNAS ESTAN POBLADAS.

  validar_modelo        valida la declaracion. No importa openpyxl siquiera
  verificar_faseA       mira la hoja, pero comprueba estructura y tipos
  verificar_documentos  compara la prosa con el modelo
  verificar_enlaces     resuelve enlaces
  verificar_reproducible  compara dos pasadas del generador entre si

Que se colo por ese hueco
-------------------------
El 2026-08-10, ocho cambios pasaron los cinco en verde y tres eran defectuosos:

  - Ubicacion_LatLong quedo vacia en las 368 filas siendo obligatoria, y es la
    columna que RG-01 desreferencia para el geofencing. DISTANCE() contra
    blanco no da error: da un valor que rechaza el cierre legitimo.
  - SED_Sedes.UnidadFuncionalID quedo vacia en 5 de 6 siendo obligatoria, y
    RG-34 la compara. La regla volvia la fila imposible de guardar.
  - ACT_Activos.SedeID nacio vacia en las 368, asi que el cambio que la
    introdujo no lo ejercita ni una fila.

Los tres son el mismo fallo: **estructura entregada sin poblacion**. Y los tres
son invisibles para un verificador que solo lea declaraciones.

La otra mitad
-------------
Una Ref que resuelve puede apuntar a lo que no es (R-04), pero una Ref que NO
resuelve es directamente una fila que la aplicacion descarta sin avisar. Eso
tampoco lo miraba nadie: verificar_app compara el NUMERO de filas, no su
contenido.

Uso:  python scripts/verificar_datos.py ["BD/<archivo>.xlsx"]
Sale con codigo 1 si una columna obligatoria esta vacia en filas que existen,
o si una referencia no resuelve.
"""
import os
import sys

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "scripts"))

try:
    import openpyxl
except ImportError:
    print("Falta openpyxl."); sys.exit(2)

from modelo_objetivo import MODELO
from alcance_reglas import por_columna
from lectura_de_vuelta import VOLCADO_CIEGO_A
from sistema import VOLCADO

ARCHIVO = sys.argv[1] if len(sys.argv) > 1 else VOLCADO
if not os.path.isabs(ARCHIVO):
    ARCHIVO = os.path.join(RAIZ, ARCHIVO)

# Columnas que el modelo declara obligatorias y que HOY se sabe que operacion
# tiene que rellenar. No son un permiso indefinido: cada una lleva la fecha a
# partir de la cual el aviso pasa a fallo, igual que hace D-04 con los
# aplazamientos. Un aviso que no caduca deja de leerse (CLAUDE.md 7.11).
POBLA_OPERACION = {
    # ACT_Activos.SedeID estuvo aqui con limite None -el aviso que no caduca
    # que este mismo comentario prohibe- y ADEMAS no se evaluaba nunca, porque
    # la columna no es obligatoria. Una excusa preinstalada: el dia que se
    # declarara obligatoria habria silenciado el fallo para siempre. Retirada.
    ("SED_Sedes", "UnidadFuncionalID"):
        ("la UF de cada edificacion sale de su PR, y el PR de las sedes lo "
         "tiene operacion, no el contrato", "2026-08-31"),
    }


def texto(v):
    return "" if v is None else str(v).strip()


wb = openpyxl.load_workbook(ARCHIVO, data_only=True, read_only=True)
datos = {}
for t in MODELO:
    if t not in wb.sheetnames:
        continue
    ws = wb[t]
    cab = [texto(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    filas = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r is None or all(v in (None, "") for v in r):
            continue
        filas.append({c: (r[i] if i < len(r) else None)
                      for i, c in enumerate(cab) if c})
    datos[t] = filas

fallos, avisos = [], []
ancho = "=" * 78
print(ancho)
print("LOS DATOS SOSTIENEN EL MODELO")
print(ancho)
try:                                    # en Windows relpath revienta entre unidades
    _mostrado = os.path.relpath(ARCHIVO, RAIZ)
except ValueError:
    _mostrado = ARCHIVO
print("Archivo: %s" % _mostrado)
print("")

# ------------------------------------------- G-01  obligatoria y sin poblar
print("G-01  Columnas obligatorias · G-04  tablas vacias, tipadas a ciegas")
print("")
# Una tabla vacia NO es "nada que medir": es donde se concentra el fallo.
#
# Este bucle hacia `continue` sobre ella, y con eso las 80 columnas de las ocho
# tablas de movimiento eran estructuralmente invisibles para el verificador. No
# porque estuvieran bien: porque no se miraban. Es la misma confusion que
# auditar_cableado.py habia escrito ese mismo dia -«confundir no lo puedo ver
# con esta bien»- aplicada a las referencias y no generalizada a los tipos.
#
# No cambia el criterio de G-01 -que mide la HOJA, y el tipo vive en el EDITOR-:
# lo que cambia es que el silencio pase a ser una obligacion contada y con
# fecha. Endurecer, no relajar.
LIMITE_TIPOS_A_CIEGAS = "2026-08-31"

for t in MODELO:
    filas = datos.get(t)
    if not filas:
        cuantas = len(MODELO[t]["columnas"])
        import datetime
        vencido = datetime.date(*map(int, LIMITE_TIPOS_A_CIEGAS.split("-")))             < datetime.date.today()
        linea = ("%s llego VACIA, asi que AppSheet eligio el tipo de sus %d "
                 "columnas sin un solo dato. Ninguna esta confirmada en el "
                 "editor" % (t, cuantas))
        if t in VOLCADO_CIEGO_A:
            linea += (". Y OJO: este archivo la vacia POR DISENO, asi que aqui "
                      "saldra vacia aunque la aplicacion tenga filas. Para "
                      "mirarla de verdad, instantanea.py")
        (fallos if vencido else avisos).append(
            "[G-04] %s%s" % (linea, ". El plazo vencio el %s" % LIMITE_TIPOS_A_CIEGAS
                             if vencido else " (hasta el %s)" % LIMITE_TIPOS_A_CIEGAS))
        continue
    for col in MODELO[t]["columnas"]:
        if not col.get("obligatoria"):
            continue
        n = col["nombre"]
        vacias = [f for f in filas if not texto(f.get(n))]
        if not vacias:
            continue
        clave = (t, n)
        if clave in POBLA_OPERACION:
            motivo, limite = POBLA_OPERACION[clave]
            import datetime
            caducado = limite and datetime.date(*map(int, limite.split("-"))) \
                < datetime.date.today()
            linea = ("%s.%s vacia en %d de %d. La rellena operacion: %s"
                     % (t, n, len(vacias), len(filas), motivo))
            if caducado:
                fallos.append("[G-01] %s. El plazo vencio el %s" % (linea, limite))
            else:
                avisos.append("[G-01] %s%s"
                              % (linea, " (hasta el %s)" % limite if limite else ""))
            continue
        fallos.append(
            "[G-01] %s.%s es OBLIGATORIA y esta vacia en %d de %d filas. La "
            "aplicacion no dejara guardar, o peor: la expresion que la lea "
            "comparara contra blanco y no dara error"
            % (t, n, len(vacias), len(filas)))

# ---------------------------------------------- G-02  referencias que no resuelven
print("G-02  Las %d referencias, contra los datos reales"
      % sum(1 for t in MODELO for c in MODELO[t]["columnas"] if c.get("ref")))
print("")
huerfanas = 0
for t in MODELO:
    for col in MODELO[t]["columnas"]:
        destino = col.get("ref")
        if not destino or destino not in MODELO:
            continue
        pk = next((c["nombre"] for c in MODELO[destino]["columnas"] if c.get("pk")), None)
        if not pk:
            continue
        validas = {texto(f.get(pk)) for f in datos.get(destino, [])}
        if not validas:
            continue    # destino vacio: no se puede juzgar, F-11 ya lo dice
        malas = [texto(f.get(col["nombre"])) for f in datos.get(t, [])
                 if texto(f.get(col["nombre"]))
                 and texto(f.get(col["nombre"])) not in validas]
        if malas:
            huerfanas += len(malas)
            ejemplo = " · ".join(sorted(set(malas))[:4])
            fallos.append(
                "[G-02] %s.%s: %d valores no existen en %s.%s. AppSheet no da "
                "error: descarta la fila. Ejemplos: %s"
                % (t, col["nombre"], len(malas), destino, pk, ejemplo))

# ------------------------------------------- G-03  una columna, un tipo Python
#
# AppSheet infiere el tipo del CONTENIDO. Una columna con 'TRUE' de cadena y
# True booleano mezclados le da una senal contradictoria, y la resuelve por
# mayoria: la minoria se pierde. Es el mismo mecanismo que descartaba a un
# tecnico entero por tener la clave alfanumerica entre diez numericas.
print("G-03  Homogeneidad de tipo dentro de cada columna")
print("")
for t in MODELO:
    for col in MODELO[t]["columnas"]:
        n = col["nombre"]
        tipos = {type(f.get(n)).__name__ for f in datos.get(t, [])
                 if f.get(n) not in (None, "")}
        if len(tipos) > 1:
            avisos.append(
                "[G-03] %s.%s mezcla %s. AppSheet tipa por la mayoria"
                % (t, n, " y ".join(sorted(tipos))))

# ------------------------- G-05  una regla que lee una columna vacia no hace nada
#
# El fallo que se repitio tres veces el 2026-08-10, y que ningun verificador
# veia porque G-01 solo mira las columnas `obligatoria`:
#
#   RG-03  bien escrita sobre una columna que AppSheet tipo Text
#   RG-06  [EstadoActivoID].[GeneraAlerta] = TRUE, y GeneraAlerta esta VACIA
#          en las 4 filas del catalogo: el bot no se dispara nunca
#   RG-19  compara Precision_GPS, que nadie puebla porque la funcion que la
#          poblaria no existe en AppSheet
#
# Las tres estan CONFIGURADAS. Ninguna da error. Ninguna hace nada.
#
# Una columna sin datos no es sospechosa por si misma -media plantilla esta
# vacia a proposito, la rellena operacion-. Lo que la vuelve grave es que **una
# regla dependa de ella**: ahi el silencio deja de ser una espera y pasa a ser
# una salvaguarda apagada.
print("G-05  Columnas de las que depende una regla, y estan vacias")
print("")
_reglas = por_columna()
for (t, n), ids in sorted(_reglas.items()):
    filas = datos.get(t)
    if not filas:
        continue      # tabla vacia: ya lo cuenta G-04, y ahi es esperado
    if not any(c["nombre"] == n for c in MODELO[t]["columnas"]):
        continue
    llenas = sum(1 for f in filas if texto(f.get(n)))
    if llenas:
        continue
    avisos.append(
        "[G-05] %s.%s esta vacia en las %d filas y de ella depende %s. "
        "La regla queda configurada y sin efecto: no da error, no hace nada"
        % (t, n, len(filas), ", ".join(sorted(ids))))

# -------------------------------------------------------------------- salida
for f in fallos:
    print("  x %s" % f)
if fallos:
    print("")
for a in avisos:
    print("  ! %s" % a)

print("")
print(ancho)
if fallos:
    print("LOS DATOS NO SOSTIENEN EL MODELO: %d fallos, %d avisos"
          % (len(fallos), len(avisos)))
    print("Estructura entregada sin poblacion. Es lo que los otros cinco")
    print("verificadores no pueden ver, porque no abren el archivo de datos.")
    print(ancho)
    sys.exit(1)

print("DATOS COHERENTES: 0 obligatorias vacias sin motivo · 0 referencias huerfanas")
print("%d avisos" % len(avisos))
print(ancho)
