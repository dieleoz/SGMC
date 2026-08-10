# -*- coding: utf-8 -*-
"""Genera BD/Modelo_Datos_PLANTILLA.xlsx entero, en un comando.

Es EL ENTREGABLE. El funcional recibe este archivo, lo completa con el dato
real y desde ahi sigue las guias. Por eso tiene que salir generado y limpio:
28 pestanas con exactamente las columnas que el modelo declara, sin las 47 que
la hoja heredada arrastraba.

Por que existe
--------------
La plantilla se armaba con dos scripts y varios pasos a mano. generar_hoja_limpia
daba la estructura, generar_inventario daba los 355, y unirlos, anadir _LEEME y
poblar el radio no estaba escrito en ningun sitio. Eso la convertia en un
artefacto que se conserva en vez de generarse -justo lo que este proyecto
decidio no volver a tener- y arrastraba dos defectos que nadie podia ver:

  - TIP_TiposActivo con 18 tipos contra 18 familias del Plan Maestro que NO son
    la misma lista. 78 activos con el checklist de otro equipo.
  - RadioGeofencingKm vacio, que hace que RG-01 compare contra blanco.

Que hace
--------
  1. Crea las 28 pestanas con las columnas de MODELO, en su orden.
  2. Migra los datos del libro origen emparejando por NOMBRE de columna.
     Lo que el modelo no declara, no viaja.
  3. Rehace TIP_TiposActivo desde catalogo_tipos: 27 tipos con su radio.
  4. Completa FRM_Formularios con los formularios que falten, uno por tipo.
  5. Anade los 355 activos sinteticos detras de los 34 del fixture.
  6. Pone _LEEME delante, que es lo unico que el funcional lee sin que se lo
     expliquen.

Uso:  python scripts/generar_plantilla.py ["BD/<origen>.xlsx"]
"""
import os
import re
import sys

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "scripts"))

from modelo_objetivo import MODELO
from catalogo_tipos import TIPOS_ACTIVO, FAMILIAS, comprobar
from generar_inventario import generar_filas, LARGO_KM, UNIDADES, unidad_funcional
from banco_preguntas import (preguntas_de, comprobar as comprobar_banco,
                             MARCA, VALORES_ESTADO)

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    print("Falta openpyxl."); sys.exit(2)

ORIGEN = sys.argv[1] if len(sys.argv) > 1 else "BD/Modelo_Datos_PLANTILLA.xlsx"
if not os.path.isabs(ORIGEN):
    ORIGEN = os.path.join(RAIZ, ORIGEN)
SALIDA = os.path.join(RAIZ, "BD", "Modelo_Datos_PLANTILLA.xlsx")

# El catalogo se comprueba ANTES de escribir nada. Si dos familias comparten
# tipo, la plantilla saldria con activos viendo el checklist de otro equipo, y
# eso no lo detecta ningun verificador posterior: la referencia resuelve.
fallos = comprobar()
if fallos:
    print("El catalogo de tipos no es coherente. No se genera nada:")
    for f in fallos:
        print("   x", f)
    sys.exit(1)


def texto(v):
    return "" if v is None else str(v).strip()


# ----------------------------------------------------------------- 1. estructura
src = openpyxl.load_workbook(ORIGEN, data_only=True, read_only=True)
dst = openpyxl.Workbook()
dst.remove(dst.active)

resumen = []
descartadas = []

for tabla in MODELO:
    cols = [c["nombre"] for c in MODELO[tabla]["columnas"]]
    ws = dst.create_sheet(tabla)
    ws.append(cols)

    if tabla not in src.sheetnames:
        resumen.append((tabla, len(cols), 0, "pestana nueva, sin datos de origen"))
        continue

    o = src[tabla]
    cab = [texto(c.value) for c in next(o.iter_rows(min_row=1, max_row=1))]
    idx = {n: i for i, n in enumerate(cab) if n}

    fuera = [n for n in cab if n and n not in cols]
    if fuera:
        descartadas.append((tabla, fuera))

    n = 0
    for r in o.iter_rows(min_row=2, values_only=True):
        if not r or all(v in (None, "") for v in r):
            continue
        fila = [r[idx[c]] if c in idx and idx[c] < len(r) else None for c in cols]
        if all(v in (None, "") for v in fila):
            continue
        ws.append(fila)
        n += 1
    resumen.append((tabla, len(cols), n, ""))

resumen = {t: [c, f, nota] for t, c, f, nota in resumen}


def columnas(tabla):
    return [c["nombre"] for c in MODELO[tabla]["columnas"]]


def escribir(tabla, filas_dict):
    """Reescribe una pestana entera desde diccionarios columna -> valor."""
    ws = dst[tabla]
    cols = columnas(tabla)
    ws.delete_rows(2, ws.max_row)
    for f in filas_dict:
        ws.append([f.get(c, "") for c in cols])
    resumen[tabla][1] = len(filas_dict)


def leer(tabla):
    """Las filas de una pestana ya construida, como diccionarios."""
    ws = dst[tabla]
    cols = columnas(tabla)
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r is None or all(v in (None, "") for v in r):
            continue
        out.append({c: (r[i] if i < len(r) else None) for i, c in enumerate(cols)})
    return out


# --------------------------------------------- 2. TIP_TiposActivo, desde el catalogo
#
# Se rehace entera, no se parchea. Los 18 que venian de la hoja traian el radio
# vacio en los 18 y el nombre de la subestacion con la tilde rota.
tipos = []
for tid, _clave, nombre, categoria, qr, gps, formulario, radio in TIPOS_ACTIVO:
    tipos.append({
        "TipoActivoID": tid,
        "Nombre": nombre,
        "Categoria": categoria,
        "FormularioID": formulario,
        "TieneQR": "TRUE" if qr else "FALSE",
        "RequiereGPS": "TRUE" if gps else "FALSE",
        "RadioGeofencingKm": radio,
        "Activo": "TRUE",
    })
escribir("TIP_TiposActivo", tipos)

# ------------------------------------------- 3. FRM_Formularios, uno por tipo
#
# Los existentes se conservan tal cual: su columna Version la leen los
# checklists ya creados (RG-09), y reescribirla dejaria el historico apuntando
# a una version de formulario que nunca existio. Solo se anaden los que faltan.
formularios = leer("FRM_Formularios")
existentes = {texto(f["FormularioID"]) for f in formularios}
anadidos = []
for tid, _clave, nombre, _cat, _qr, _gps, formulario, _radio in TIPOS_ACTIVO:
    if formulario in existentes:
        continue
    formularios.append({
        "FormularioID": formulario,
        "Nombre": "Checklist %s" % nombre,
        "Descripcion": "Checklist mantenimiento preventivo %s" % nombre,
        "Version": 1,
        "Activo": "TRUE",
    })
    anadidos.append(formulario)
escribir("FRM_Formularios", formularios)

# --------------------------------------------------- 4. un solo inventario
#
# Antes habia dos listas conviviendo: los 34 activos que venian de la hoja, con
# codigo SOS-001, y 355 generados con codigo SOS_1. Dos convenciones para lo
# mismo, y las dos con un SOS numero 1.
#
# Ahora es una sola. Se completa cada familia hasta la cantidad del Plan Maestro
# contando lo que ya hay, con el formato de operacion. Los codigos del fixture se
# normalizan al prefijo de su familia -SVR pasa a SERV- y se renumeran 001, 002,
# 003 dentro de ella.
#
# **13 de los 34 no pertenecen a ninguna familia del Plan Maestro** y hay que
# conservarlos: la fibra, el generador, el video wall, el router, el firewall, la
# UPS, el NAS, la subestacion y la bascula estatica. El Plan Maestro los cuenta en
# km, en global o en meses, no en unidades. Si se borraran, nueve de los 27 tipos
# se quedarian sin un solo activo y su checklist no se podria probar con nada.
ALIAS_FAMILIA = {"SVR": "SERV"}     # el fixture llamaba SVR al servidor

activos = leer("ACT_Activos")
prefijos_familia = {f[0] for f in FAMILIAS}
cuenta = {}
solo_fixture = 0
for a in activos:
    cod = texto(a["CodigoActivo"])
    pref = re.split(r"[-_]", cod)[0]
    pref = ALIAS_FAMILIA.get(pref, pref)
    if pref in prefijos_familia:
        cuenta[pref] = cuenta.get(pref, 0) + 1
        a["CodigoActivo"] = "%s-%03d" % (pref, cuenta[pref])
    else:
        solo_fixture += 1

fixture = len(activos)
for f in generar_filas(cuenta):
    activos.append({c: f.get(c, "") for c in columnas("ACT_Activos")})
# La unidad funcional se DERIVA del PR, no se arrastra. Venia de un reparto en
# cuartos iguales que no existe: con los tramos reales de la ANI, la frontera
# entre la primera y la segunda esta en el km 49 y no en el 34, asi que decenas
# de activos estaban en la UF equivocada. Y la UF es lo que decide que ve cada
# tecnico: RG-04 filtra por ella.
recolocados = 0
for a in activos:
    m = re.match(r"(\d+)\+(\d+)", texto(a.get("PK")) or texto(a.get("PR")))
    if not m:
        continue
    km = int(m.group(1)) + int(m.group(2)) / 1000.0
    correcta = str(unidad_funcional(min(km / LARGO_KM, 0.999999)))
    if texto(a.get("UnidadFuncionalID")) != correcta:
        a["UnidadFuncionalID"] = correcta
        recolocados += 1
    # El valor que traiamos en PR era en realidad el kilometro lineal: un PK con
    # la etiqueta equivocada. Se muda a su columna y PR queda vacio, que es la
    # verdad -el PR de INVIAS no lo sabemos y no se inventa-.
    if not texto(a.get("PK")) and texto(a.get("PR")):
        a["PK"] = texto(a.get("PR"))
        a["PR"] = ""
    elif texto(a.get("PR")) and texto(a.get("PR")) == texto(a.get("PK")):
        # Mismo valor en las dos: es la copia con la etiqueta equivocada de una
        # pasada anterior. El PK se queda; el PR se vacia, porque el de INVIAS
        # no lo sabemos.
        a["PR"] = ""

escribir("ACT_Activos", activos)

# ------------------------------- 4a. claves que AppSheet descartaria en silencio
#
# AppSheet TIPA LA COLUMNA CLAVE SEGUN LA MAYORIA de sus valores. En
# USR_Usuarios habia diez claves numericas y una alfanumerica -3aa202ee, que
# AppSheet mismo genero con UNIQUEID al crear la fila desde la aplicacion-, asi
# que tipaba la columna como Number y DESCARTABA esa fila. Sin error y sin
# aviso: Diego Escobar, tecnico, simplemente no existia para la aplicacion. Lo
# vimos porque la API devolvia 10 usuarios y la hoja tenia 11.
#
# Se renumera al siguiente libre. Va explicito y no derivado a proposito: una
# clave se cambia a mano y con constancia, nunca por una regla automatica que
# adivine, porque si algo la referenciara la reescritura silenciosa seria peor
# que el problema. Se comprobo antes que nadie la referencia.
CLAVES_RENUMERADAS = {
    # Un tecnico que la aplicacion no veia.
    "USR_Usuarios": {"3aa202ee": "12"},
    # Y al reves: aqui la mayoria es texto -las 104 generadas- y las cuatro
    # numericas son las del banco de SOS, el unico acordado. AppSheet tiparia
    # Text y las descartaria, dejando el desplegable de la primera pregunta del
    # checklist de SOS vacio. Pasan a la convencion de las demas.
    "LST_ValoresLista": {"1": "SOS001-1", "2": "SOS001-2",
                         "3": "SOS001-3", "4": "SOS001-4"},
    }

renumeradas = []
for tabla, mapa in CLAVES_RENUMERADAS.items():
    filas = leer(tabla)
    clave = columnas(tabla)[0]
    for f in filas:
        if texto(f[clave]) in mapa:
            nueva = mapa[texto(f[clave])]
            renumeradas.append("%s: %s -> %s" % (tabla, texto(f[clave]), nueva))
            f[clave] = nueva
    escribir(tabla, filas)

# ------------------------------------- 4b. valores de catalogo que el dominio exige
#
# CAL_Calzadas traia solo Izquierda y Derecha. Operacion usa tres costados: un
# equipo puede estar a la izquierda, a la derecha o EN EL CENTRO, y en doble
# calzada ese centro se llama separador. Sin la tercera fila, quien instale un
# panel en el separador tiene que elegir un costado que no es el suyo, o dejarlo
# en blanco: las dos opciones ensucian el dato y ninguna da error.
CATALOGO_MINIMO = {
    # Las seis edificaciones reales. Las cuatro filas UF1 a UF4 que tenia esta
    # tabla NO vuelven: eran unidades funcionales metidas en el catalogo de
    # sedes, con la misma clave y el mismo nombre que UNF_UnidadesFuncionales.
    #
    # Solo el peaje de Macheta trae ubicacion, y sale del contrato -Apendice
    # Tecnico 1, Tabla 2, pagina 5: PR27+240 de la Ruta 5607, cobro
    # bidireccional, UF1-. De las otras cinco el contrato no da abscisa en las
    # paginas leidas, asi que se dejan en blanco para que las ponga operacion.
    # Inventarlas seria peor que dejarlas vacias: el equipo bajo techo hereda de
    # aqui donde esta.
    "SED_Sedes": [
        {"SedeID": "1", "Nombre": "CCO", "Ciudad": "Sutatenza",
         "UnidadFuncionalID": "", "PR": "", "Ubicacion_LatLong": "", "Activo": "TRUE"},
        {"SedeID": "2", "Nombre": "Bogota", "Ciudad": "Bogota",
         "UnidadFuncionalID": "", "PR": "", "Ubicacion_LatLong": "", "Activo": "TRUE"},
        {"SedeID": "3", "Nombre": "Peaje Macheta", "Ciudad": "Macheta",
         "UnidadFuncionalID": "7", "PR": "27+240", "TramoINVIAS": "5607",
         "PK": "", "Ubicacion_LatLong": "", "Activo": "TRUE"},
        {"SedeID": "4", "Nombre": "Peaje SLG", "Ciudad": "San Luis de Gaceno",
         "UnidadFuncionalID": "", "PR": "", "Ubicacion_LatLong": "", "Activo": "TRUE"},
        {"SedeID": "5", "Nombre": "Bascula Macheta", "Ciudad": "Macheta",
         "UnidadFuncionalID": "", "PR": "", "Ubicacion_LatLong": "", "Activo": "TRUE"},
        {"SedeID": "6", "Nombre": "Bascula SLG", "Ciudad": "San Luis de Gaceno",
         "UnidadFuncionalID": "", "PR": "", "Ubicacion_LatLong": "", "Activo": "TRUE"},
        ],
    "CAL_Calzadas": [
        {"CalzadaID": "3", "Nombre": "Separador",
         "Activo": "TRUE"},
        ],
    }

# SE EMPAREJA POR NOMBRE, NO POR CLAVE. La clave es un identificador generado y
# la resiembra la cambia -SED-001 hoy, otra cosa manana-, asi que buscar por ella
# no encuentra nada en la siguiente pasada y el catalogo vuelve a anadir las
# mismas filas. Paso: SED_Sedes acabo con las seis edificaciones DUPLICADAS y
# CAL_Calzadas con dos Separador, y cada ejecucion habria anadido seis mas.
#
# El nombre es la clave natural: es lo que identifica la fila para una persona y
# lo unico que no cambia cuando se resiembran los identificadores.
anadidos_catalogo, completados_catalogo, duplicadas_catalogo = [], [], []
for tabla, filas_min in CATALOGO_MINIMO.items():
    actuales = leer(tabla)
    clave = columnas(tabla)[0]
    natural = "Nombre" if "Nombre" in columnas(tabla) else clave

    # Primero se limpia lo que la version anterior duplico.
    vistas, limpias = set(), []
    for f in actuales:
        k = texto(f.get(natural))
        if k and k in vistas:
            duplicadas_catalogo.append("%s.%s" % (tabla, k))
            continue
        vistas.add(k)
        limpias.append(f)
    actuales = limpias
    tiene = {texto(f.get(natural)) for f in actuales}
    por_clave = {texto(f.get(natural)): f for f in actuales}
    for fila in filas_min:
        k = texto(fila.get(natural))
        if k not in tiene:
            actuales.append(fila)
            anadidos_catalogo.append("%s.%s" % (tabla, k))
            continue
        # La fila ya esta: se COMPLETA lo que este vacio y NUNCA se pisa lo que
        # alguien haya escrito. Antes solo se garantizaba la existencia, y el
        # tramo del peaje de Macheta -que sale del contrato- no llego nunca a la
        # fila porque la fila ya existia sin el.
        actual = por_clave[k]
        for col, valor in fila.items():
            if col == clave:
                continue          # la clave la pone la resiembra, no el catalogo
            if valor not in (None, "") and not texto(actual.get(col)):
                actual[col] = valor
                completados_catalogo.append("%s.%s.%s" % (tabla, k, col))
    escribir(tabla, actuales)

# ------------------------------------------- 5. los bancos de preguntas
#
# Tres bancos ya existian con 15 preguntas cada uno, pero DOS de ellos vivian en
# pestanas retiradas -FRM_CCTV y FRM_PMVF- que no viajan a la plantilla. Estaban
# escritos y acordados, y se estaban perdiendo: FRM_Preguntas solo traia el de
# SOS. Esos dos se RECUPERAN tal cual, sin marca de borrador.
#
# Los demas se generan como borrador desde banco_preguntas, y cada pregunta lo
# dice de si misma en su ayuda. El funcional corrige; no parte de una hoja en
# blanco ni tiene que adivinar el formato.
fallos_banco = comprobar_banco()
if fallos_banco:
    print("El borrador de checklist no es coherente. No se genera nada:")
    for f in fallos_banco:
        print("   x", f)
    sys.exit(1)

COLS_PREG = columnas("FRM_Preguntas")
preguntas = leer("FRM_Preguntas")
con_banco = {texto(p["FormularioID"]) for p in preguntas}
valores = leer("LST_ValoresLista")
recuperados, generados = [], []


def fila_pregunta(pid, formulario, orden, sec, texto_p, tipo, obligatoria,
                  mini, maxi, unidad, ayuda):
    return {
        "PreguntaID": pid, "FormularioID": formulario, "SeccionID": sec,
        "Orden": orden, "Pregunta": texto_p, "TipoRespuestaID": tipo,
        "Obligatoria": "TRUE" if obligatoria else "FALSE",
        "ValorMinimo": mini, "ValorMaximo": maxi, "Unidad": unidad,
        "Ayuda": ayuda, "VisibleSi": "", "RequiereFoto": "FALSE",
        "Version": 1, "RequiereGPS": "FALSE", "RequiereFirma": "FALSE",
        "Activo": "TRUE",
    }


def anadir_valores(pid):
    """Los cuatro valores del desplegable de estado.

    Una pregunta de tipo Lista sin valores le muestra al tecnico un desplegable
    vacio, y eso no da error en ninguna parte.
    """
    for i, v in enumerate(VALORES_ESTADO, 1):
        valores.append({"ValorListaID": "%s-%d" % (pid, i), "PreguntaID": pid,
                        "Valor": v, "Orden": i, "Activo": "TRUE"})


# --- los dos bancos reales que estaban en pestanas retiradas
for pestana, formulario in (("FRM_CCTV", "FRM_CCTV"), ("FRM_PMVF", "FRM_PMVF")):
    if pestana not in src.sheetnames or formulario in con_banco:
        continue
    o = src[pestana]
    cab = [texto(c.value) for c in next(o.iter_rows(min_row=1, max_row=1))]
    # el encabezado de estas dos trae 'Seccion' con la tilde rota por codificacion
    idx = {n: i for i, n in enumerate(cab) if n}
    col_sec = next((n for n in cab if n.startswith("Secci")), None)
    clave = next(t[1] for t in TIPOS_ACTIVO if t[6] == formulario)
    orden = 0
    for r in o.iter_rows(min_row=2, values_only=True):
        if not r or r[0] in (None, ""):
            continue
        orden += 1
        pid = "%s%03d" % (clave, orden)
        tipo = r[idx["TipoRespuestaID"]]
        preguntas.append(fila_pregunta(
            pid, formulario, orden, r[idx[col_sec]], r[idx["Pregunta"]], tipo,
            texto(r[idx["Obligatoria"]]).upper() == "TRUE",
            r[idx["ValorMinimo"]] or "", r[idx["ValorMaximo"]] or "",
            r[idx["Unidad"]] or "", r[idx["Ayuda"]] or ""))
        if str(tipo) == "2":
            anadir_valores(pid)
    con_banco.add(formulario)
    recuperados.append("%s (%d)" % (formulario, orden))

# --- borrador para los que siguen sin banco
for tid, clave, nombre, categoria, _qr, _gps, formulario, _radio in TIPOS_ACTIVO:
    if formulario in con_banco:
        continue
    for orden, (sec, texto_p, tipo, obl, mini, maxi, unidad, ayuda) in \
            enumerate(preguntas_de(categoria), 1):
        pid = "%s%03d" % (clave, orden)
        preguntas.append(fila_pregunta(
            pid, formulario, orden, sec, texto_p, tipo, obl, mini, maxi, unidad,
            "%s %s" % (ayuda, MARCA)))
        if tipo == 2:
            anadir_valores(pid)
    generados.append(formulario)

escribir("FRM_Preguntas", preguntas)
escribir("LST_ValoresLista", valores)

# ------------------------------------- 6. los tramos de las unidades funcionales
#
# Estaban las cuatro vacias, y sin tramo el filtro por zona no tiene de que
# colgar. Se reparten los 137,03 km oficiales entre las cuatro, que es lo que
# hace el generador de inventario para decidir a que UF cae cada activo. Es un
# reparto uniforme y por tanto provisional: operacion tiene los PR reales.
# Las cuatro con su nombre y su tramo REALES, de la ANI. Antes se repartia el
# corredor en cuartos iguales, que no lo son: la primera mide 49 km y la tercera
# 18. Con el reparto uniforme, un activo del kilometro 60 caia en la UF2 cuando
# esta en la UF2 de verdad por poco, y uno del 80 caia en la UF3 cuando esta en
# la UF3 tambien por poco -pero los limites estaban a 34, 68 y 103 km, que no
# son ninguna frontera real-.
def pk(km):
    return "%02d+%03d" % (int(km), int(round((km - int(km)) * 1000)))


escribir("UNF_UnidadesFuncionales", [
    {"UnidadFuncionalID": str(uid), "Nombre": nombre,
     "PKInicial": pk(desde), "PKFinal": pk(hasta),
     "PRInicial": pri, "PRFinal": prf, "Activo": "TRUE"}
    for uid, nombre, desde, hasta, pri, prf in UNIDADES])

# ------------------------------------------- 7. fuera los registros de prueba
#
# La plantilla es lo que recibe el funcional, y las filas de prueba de otro no
# son dato suyo: son ruido que tiene que distinguir y borrar. Se van las ocho
# tablas de movimiento y quedan los catalogos, los usuarios, las asignaciones y
# el inventario, que es lo que hay que completar.
#
# Se llevan por delante los fixtures de P-08 y P-09. No es perdida: una fila de
# mantenimiento escrita a mano en la hoja no prueba que la aplicacion sepa
# crearla. Esos dos registros hay que rehacerlos DESDE la aplicacion, que es lo
# que la prueba mide en realidad.
DE_PRUEBA = ["OT_OrdenesTrabajo", "MAN_Mantenimientos", "CHK_Checklists",
             "CHD_ChecklistDetalle", "FOT_Fotografias", "FIR_Firmas",
             "NOV_Novedades", "PLA_PlanMantenimiento"]
retiradas_prueba = {}
for tabla in DE_PRUEBA:
    n = len(leer(tabla))
    if n:
        retiradas_prueba[tabla] = n
    escribir(tabla, [])

# ------------------------------------------------------------------ 8. _LEEME
#
# Va delante porque es lo unico que el funcional lee sin que se lo expliquen.
sin_preguntas = sorted({texto(f["FormularioID"]) for f in formularios}
                       - {texto(p["FormularioID"]) for p in leer("FRM_Preguntas")})

borrador = sum(1 for x in preguntas if MARCA in texto(x["Ayuda"]))
acordadas = len(preguntas) - borrador
form_cubiertos = {texto(x["FormularioID"]) for x in preguntas}

# Los formatos se derivan de los tipos que declara el modelo, con un ejemplo
# sacado del propio archivo. Es la pregunta que hace todo el que abre la hoja
# -en que formato va la fecha, como se escribe una coordenada- y hasta ahora
# habia que adivinarla.
EJEMPLO = {
    "LatLong":         ("4.812345, -73.201234", "latitud, longitud. Punto decimal, seis decimales"),
    "Date":            ("2026-08-07", "ano-mes-dia. Escribalo como fecha, no como texto"),
    "DateTime":        ("2026-08-07 14:30", "ano-mes-dia hora:minuto, 24 horas"),
    "ChangeTimestamp": ("(no se escribe)", "la pone el servidor al guardar. Dejela vacia"),
    "Yes/No":          ("TRUE / FALSE", "en mayusculas, sin comillas"),
    "Decimal":         ("0,05", "coma decimal"),
    "Number":          ("12", "entero, sin separador de miles"),
    "Enum":            ("", "uno de los valores permitidos, escrito igual"),
    "Ref":             ("", "el identificador de la fila destino, tal cual"),
    }

formatos = []
for tabla in MODELO:
    for c in MODELO[tabla]["columnas"]:
        if c["tipo"] in EJEMPLO and c["tipo"] not in [f[0] for f in formatos]:
            formatos.append((c["tipo"], EJEMPLO[c["tipo"]][0], EJEMPLO[c["tipo"]][1]))

LEEME = [
    ("PLANTILLA DE DATOS - SGMC", True),
    ("", False),
    ("%d pestanas de datos mas esta, con la estructura exacta que espera la aplicacion." % len(MODELO), False),
    ("NO anada ni quite columnas: la aplicacion las lee por nombre.", False),
    ("", False),
    ("Viene AUTOCOMPLETADA a proposito.", True),
    ("Cada columna trae ya un valor con el formato correcto, para que usted", False),
    ("corrija en vez de adivinar. Lo que no sea cierto, cambielo.", False),
    ("", False),
    ("NINGUNA COORDENADA DE ESTE ARCHIVO ES REAL", True),
    ("Las %d filas de ACT_Activos tienen coordenada, pero esta calculada sobre el" % len(activos), False),
    ("trazado del corredor, no levantada en campo. Cada fila lo dice en Observaciones.", False),
    ("", False),
    ("EN QUE FORMATO SE ESCRIBE CADA COSA", True),
    ("", False),
    ("TIPO | EJEMPLO | REGLA", False),
    ] + [("%s | %s | %s" % (t, e or "-", r), False) for t, e, r in formatos] + [
    ("", False),
    ("LO QUE HAY QUE COMPLETAR", True),
    ("", False),
    ("PESTANA | COLUMNA | QUE PONER", False),
    ("ACT_Activos | Ubicacion | LA COORDENADA REAL del equipo", False),
    ("    SIN ESTO NINGUN TECNICO PUEDE CERRAR UNA ORDEN EN VIA", False),
    ("ACT_Activos | PR | El punto de referencia INVIAS real", False),
    ("ACT_Activos | CodigoActivo | El codigo con el que operacion conoce el equipo,", False),
    ("    si no coincide con el que trae", False),
    ("ACT_Activos | Criticidad | Alta / Media / Baja. Vacia en las %d" % len(activos), False),
    ("FRM_Preguntas | Pregunta y Ayuda | LAS %d QUE LLEVAN [BORRADOR] SON PROPUESTA NUESTRA." % borrador, False),
    ("    Corrijalas y quite la marca. Buscar BORRADOR en la hoja dice que falta.", False),
    ("    Las otras %d ya estaban acordadas: SOS, CCTV y PMVF." % acordadas, False),
    ("USR_Usuarios | Correo | EXACTAMENTE la cuenta con la que inicia sesion", False),
    ("ASG_AsignacionZona | todas | Que unidad funcional atiende cada tecnico.", False),
    ("    Sin fila, ese tecnico no ve ningun activo.", False),
    ("UNF_UnidadesFuncionales | PRInicial/Final | Traen un reparto uniforme de los", False),
    ("    %s km del corredor. Ponga los tramos reales." % ("%.2f" % LARGO_KM).replace(".", ","), False),
    ("", False),
    ("LO QUE NO HAY QUE TOCAR", True),
    ("", False),
    ("Nombres de pestanas y de columnas.", False),
    ("Las columnas de identificador: relacionan las tablas entre si.", False),
    ("TIP_TiposActivo.RadioGeofencingKm: es la distancia para poder cerrar, por tipo.", False),
    ("    Dejarla en blanco hace que se rechace TAMBIEN el cierre legitimo.", False),
    ("Las columnas de tipo ChangeTimestamp: las escribe el servidor.", False),
    ("", False),
    ("COMO ESTAN LOS DATOS HOY", True),
    ("", False),
    ("ACT_Activos | %d filas | Un solo inventario, codigo SOS-001, CCTV-001, SWIT-001" % len(activos), False),
    ("    Cada familia suma lo que dice el Plan Maestro: 54 SOS, 142 switches, 26 camaras", False),
    ("    %d equipos que el Plan Maestro no cuenta por unidades -fibra, generador," % solo_fixture, False),
    ("    video wall, router, firewall, UPS, NAS, subestacion y bascula estatica-", False),
    ("TIP_TiposActivo | %d filas | Un tipo por checklist, con su radio de cierre" % len(TIPOS_ACTIVO), False),
    ("FRM_Formularios | %d filas | Uno por tipo" % len(formularios), False),
    ("FRM_Preguntas | %d preguntas | Los %d formularios tienen checklist" % (len(preguntas), len(form_cubiertos)), False),
    ("USR_Usuarios | %d filas | Personas reales" % len(leer("USR_Usuarios")), False),
    ("", False),
    ("SIN REGISTROS DE PRUEBA.", True),
    ("Se retiraron las ordenes, mantenimientos, checklists, fotografias y firmas", False),
    ("de ensayo que traia la hoja: no son dato suyo y habria que distinguirlos.", False),
    ("", False),
    ("Documentacion: repositorio del proyecto, empezando por ESTADO.md", False),
]

# ------------------------- claves alfanumericas, para que AppSheet infiera Text
#
# AppSheet infiere el tipo de una columna mirando el nombre de la cabecera Y el
# contenido de las filas. Una clave cuyos valores son 1, 2, 3 la tipa Number
# aunque el modelo la declare Text, y entonces pasan dos cosas malas: una fila
# con clave alfanumerica -las que genera UNIQUEID desde la aplicacion- se
# DESCARTA sin avisar, y las referencias comparan numero contra texto.
#
# Once tablas tenian clave numerica, heredada de la hoja vieja. Las otras diez
# del modelo ya usaban clave legible -OT-0001, MOT-01, FAL-01, ASG-01, PLA-001,
# FRM_SOS, SOS001, UMBRAL_GPS-, asi que esto no inventa una convencion: la
# completa.
#
# El cambio se PROPAGA solo: se reescribe la clave y todas las columnas Ref que
# apuntan a esa tabla, derivadas de MODELO. Hacerlo a mano seria garantizar el
# olvido de una.
SEMILLA_CLAVES = {
    "SED_Sedes":               ("SED-%03d", 1),
    "UNF_UnidadesFuncionales": ("UNF-%02d", 1),
    "ROL_Roles":               ("ROL-%02d", 1),
    "USR_Usuarios":            ("USR-%03d", 1),
    "TIP_TiposActivo":         ("TIP-%03d", 1),
    "EST_Activo":              ("EST-%02d", 1),
    "FRE_Frecuencias":         ("FRE-%02d", 1),
    "CAL_Calzadas":            ("CAL-%02d", 1),
    "ACT_Activos":             ("ACT-%04d", 1),
    "FRM_Secciones":           ("SEC-%02d", 1),
    "TPR_TiposRespuesta":      ("TPR-%02d", 1),
    }

resembradas = []
for tabla, (patron, desde) in SEMILLA_CLAVES.items():
    filas = leer(tabla)
    if not filas:
        continue
    clave = columnas(tabla)[0]
    mapa = {}
    for i, f in enumerate(filas, desde):
        viejo = texto(f[clave])
        if not viejo:
            continue
        mapa[viejo] = patron % i
        f[clave] = mapa[viejo]
    escribir(tabla, filas)
    # y ahora todo lo que apuntaba a esa tabla
    tocadas = 0
    for t2 in MODELO:
        cols_ref = [c["nombre"] for c in MODELO[t2]["columnas"] if c.get("ref") == tabla]
        if not cols_ref:
            continue
        filas2 = leer(t2)
        for f2 in filas2:
            for c2 in cols_ref:
                v = texto(f2.get(c2))
                if v in mapa:
                    f2[c2] = mapa[v]
                    tocadas += 1
        escribir(t2, filas2)
    resembradas.append("%s (%d claves, %d referencias)" % (tabla, len(mapa), tocadas))

# --------------------------------------- claves y referencias, todas a texto
#
# El modelo declara las claves como Text y las referencias apuntan a ellas. La
# hoja heredada guardaba unas filas como texto -'1'- y las generadas salian como
# numero, asi que ACT_Activos.ActivoID mezclaba los dos tipos.
#
# AppSheet compara el valor de la referencia con el de la clave. Con la clave
# mezclada la comparacion depende de como llegue cada fila, y el sintoma es una
# referencia que a veces resuelve y a veces no, sin error.
def a_texto(v):
    if v is None or v == "":
        return v
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return v


normalizadas = 0
for tabla in MODELO:
    objetivo = [i for i, c in enumerate(MODELO[tabla]["columnas"])
                if c.get("pk") or c.get("tipo") == "Ref"]
    if not objetivo:
        continue
    for fila in dst[tabla].iter_rows(min_row=2):
        for i in objetivo:
            nuevo = a_texto(fila[i].value)
            if nuevo is not fila[i].value:
                fila[i].value = nuevo
                normalizadas += 1

hoja = dst.create_sheet("_LEEME", 0)
for texto_fila, negrita in LEEME:
    hoja.append([texto_fila])
    if negrita:
        hoja.cell(row=hoja.max_row, column=1).font = Font(bold=True)
hoja.column_dimensions["A"].width = 95

dst.save(SALIDA)

# ------------------------------------------------------------------- informe
print("Generado:", SALIDA)
print("Origen:  ", os.path.basename(ORIGEN))
print()
print("%-26s %6s %8s  %s" % ("TABLA", "COLS", "FILAS", ""))
for t in sorted(resumen):
    c, f, nota = resumen[t]
    print("%-26s %6d %8d  %s" % (t, c, f, nota))
print()
print("Pestanas: %d + _LEEME   ·   Columnas: %d   ·   Filas: %d"
      % (len(resumen), sum(v[0] for v in resumen.values()),
         sum(v[1] for v in resumen.values())))
print()
print("Activos recolocados en su unidad funcional real: %d" % recolocados)
for r in resembradas:
    print("Clave resembrada: %s" % r)
print("Claves y referencias normalizadas a texto: %d" % normalizadas)
if anadidos_catalogo:
    print("Valores de catalogo anadidos: %s" % " · ".join(anadidos_catalogo))
if renumeradas:
    print("Claves renumeradas: %s" % " · ".join(renumeradas))
if duplicadas_catalogo:
    print("Filas duplicadas retiradas: %s" % " · ".join(duplicadas_catalogo))
if completados_catalogo:
    print("Campos de catalogo completados: %s" % " · ".join(completados_catalogo))
print()
print("Tipos de activo: %d   ·   Formularios anadidos: %d   ·   Activos: %d (%d fixture + %d generados)"
      % (len(TIPOS_ACTIVO), len(anadidos), len(activos), fixture, len(activos) - fixture))
if anadidos:
    print("   ", " · ".join(anadidos))
print()
if descartadas:
    total = sum(len(f) for _, f in descartadas)
    print("=== %d columnas del origen que NO viajan (el modelo no las declara) ===" % total)
    for t, fuera in sorted(descartadas):
        print("  %-24s %s" % (t, " · ".join(sorted(fuera))))
print()
print("Comprueba con:  python scripts/verificar_faseA.py \"BD/Modelo_Datos_PLANTILLA.xlsx\"")
