# -*- coding: utf-8 -*-
"""Compara lo que la APLICACION ve con lo que el repositorio dice que hay.

NO es uno de los seis verificadores, aunque se llame igual y durante un tiempo
se describiera a si mismo como «el quinto». Aquellos leen ARCHIVOS -el modelo,
un .xlsx, la prosa, los enlaces- y se corren antes de cerrar cualquier cambio.

Este pregunta a la APLICACION EN VIVO, y por eso pertenece al otro grupo, con
auditar_cableado.py y instantanea.py: instrumentos que miran produccion. La
diferencia importa porque los tres comparten un limite que los seis no tienen
-lo que la API no devuelve, no se puede ver- y porque ninguno de ellos es un
gate: no bloquean, informan.

Por que hacia falta
-------------------
El 2026-08-10, la primera consulta a la API devolvio 10 filas en USR_Usuarios
donde la plantilla tiene 11. Una fila menos, y era un tecnico: sin su fila no
inicia sesion ni recibe ordenes. Nada en el repositorio podia detectarlo, porque
todo lo que verificabamos era el archivo de origen, no lo que llego.

**Subir un archivo no es lo mismo que que la aplicacion lo lea.** Entre los dos
hay una conversion a Hoja de calculo y un escaneo de AppSheet, y los dos pueden
perder algo sin decirlo.

SOLO LECTURA. Usa unicamente la accion Find. No expone Add, Edit ni Delete: un
verificador que pueda escribir deja de ser un verificador.

Uso
---
    set APPSHEET_API_KEY=...        (o ponerla en .env, que esta en .gitignore)
    python scripts/verificar_app.py

Sale con codigo 1 si la aplicacion no ve lo que el repositorio declara.
"""
import json
import os
import sys
import urllib.error
import urllib.request

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "scripts"))

from modelo_objetivo import MODELO
from sistema import APP_ID, APP_NOMBRE, VOLCADO

try:
    import openpyxl
except ImportError:
    print("Falta openpyxl."); sys.exit(2)

# La clave se lee del entorno o de .env. NUNCA se escribe en el repositorio ni
# se imprime: este script la usa y no la muestra.
env = os.path.join(RAIZ, ".env")
if os.path.exists(env):
    for linea in open(env, encoding="utf-8"):
        linea = linea.strip()
        if linea and not linea.startswith("#") and "=" in linea:
            k, v = linea.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())

API_KEY = os.environ.get("APPSHEET_API_KEY", "")
if not API_KEY:
    print("Falta APPSHEET_API_KEY, en el entorno o en .env.")
    sys.exit(2)

URL = "https://api.appsheet.com/api/v2/apps/%s/tables/%s/Action" % (APP_ID, "%s")


def leer(tabla):
    """Las filas que la aplicacion ve en esa tabla. Solo Find."""
    payload = json.dumps({
        "Action": "Find",
        "Properties": {"Locale": "es-CO", "Timezone": "SA Pacific Standard Time"},
    }).encode("utf-8")
    req = urllib.request.Request(
        URL % tabla, data=payload,
        headers={"ApplicationAccessKey": API_KEY, "Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=60) as r:
        cuerpo = r.read().decode("utf-8")
    return json.loads(cuerpo) if cuerpo else []


# --------------------------------------------------- lo que dice el repositorio
libro = openpyxl.load_workbook(os.path.join(RAIZ, VOLCADO),
                               read_only=True, data_only=True)


def filas_plantilla(tabla):
    ws = libro[tabla]
    return sum(1 for r in ws.iter_rows(min_row=2, values_only=True)
               if r and r[0] not in (None, ""))


def claves_plantilla(tabla):
    ws = libro[tabla]
    return {str(r[0]).strip() for r in ws.iter_rows(min_row=2, values_only=True)
            if r and r[0] not in (None, "")}


fallos, avisos, oks = [], [], []
ancho = "=" * 78
print(ancho)
print("LA APLICACION CONTRA EL REPOSITORIO")
print(ancho)
print("Aplicacion: %s" % APP_NOMBRE)
print("Se compara contra: %s" % VOLCADO)
print("")
print("%-26s %8s %8s  %s" % ("TABLA", "APP", "REPO", ""))

for tabla in sorted(MODELO):
    if tabla not in libro.sheetnames:
        continue
    esperadas = filas_plantilla(tabla)
    try:
        vistas = leer(tabla)
    except urllib.error.HTTPError as e:
        fallos.append("[A-01] %s: la aplicacion no responde (HTTP %s). Si la tabla no esta "
                      "dada de alta, el cableado no puede empezar" % (tabla, e.code))
        print("%-26s %8s %8d  NO RESPONDE" % (tabla, "-", esperadas))
        continue
    except Exception as e:
        fallos.append("[A-01] %s: %s" % (tabla, e))
        print("%-26s %8s %8d  ERROR" % (tabla, "-", esperadas))
        continue

    n = len(vistas) if isinstance(vistas, list) else 0
    nota = ""
    if n != esperadas:
        # Una fila que el repositorio declara y la aplicacion no ve es un dato
        # perdido entre el archivo y el sistema. Al reves, es un dato que alguien
        # escribio en la hoja y el repositorio no conoce.
        if n < esperadas:
            faltan = claves_plantilla(tabla) - {
                str(f.get(list(f.keys())[0], "")).strip() for f in vistas} if vistas else claves_plantilla(tabla)
            fallos.append("[A-02] %s: la aplicacion ve %d filas y el repositorio declara %d. "
                          "Faltan %s" % (tabla, n, esperadas,
                                         ", ".join(sorted(faltan)[:6]) or "(no identificadas)"))
            nota = "FALTAN %d" % (esperadas - n)
        else:
            avisos.append("[A-03] %s: la aplicacion ve %d filas y el repositorio declara %d. "
                          "Alguien escribio en la hoja y el repositorio no lo sabe"
                          % (tabla, n, esperadas))
            nota = "sobran %d" % (n - esperadas)
    else:
        oks.append(tabla)
    print("%-26s %8d %8d  %s" % (tabla, n, esperadas, nota))

print("")
print(ancho)
if oks:
    print("COINCIDEN (%d)" % len(oks))
if avisos:
    print("")
    print("AVISOS (%d)" % len(avisos))
    for a in avisos:
        print("  - %s" % a)
if fallos:
    print("")
    print("FALLOS (%d)" % len(fallos))
    for f in fallos:
        print("  x %s" % f)
    print(ancho)
    print("LA APLICACION NO VE LO QUE EL REPOSITORIO DECLARA")
    print(ancho)
    sys.exit(1)

print(ancho)
print("LA APLICACION VE EXACTAMENTE LO QUE EL REPOSITORIO DECLARA")
print(ancho)
