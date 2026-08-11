# -*- coding: utf-8 -*-
"""Genera docs/MANUAL_DESPLIEGUE.md desde el modelo.

Es **el unico documento que hay que leer para desplegar de cero**. Los otros
cuatro -PROMPT_CABLEADO, PROMPT_EXPRESIONES, TIPOS_ESPERADOS y
CORRECCIONES_CABLEADO- siguen existiendo, pero como EXTRACTOS QUE SE LE PASAN A
UN EJECUTOR, no como piezas que haya que ir a buscar. El manual dice cual es
cada uno y cuando se usa; el camino esta aqui dentro.

Ninguna cifra se escribe a mano
-------------------------------
Todas salen de `modelo_objetivo.py`, de `inferencia.py`, del volcado de la hoja
o de la ultima instantanea de la aplicacion. Un numero escrito a mano en un
generador es indistinguible de uno derivado cuando se lee el .md, y envejece
igual de callado: este mismo encabezado decia «las 38 referencias y las 20
reglas» cuando el modelo ya declaraba 39 y 21.

Esta escrito por ROL, no por persona, para poder replicarlo en otro contrato.

Uso:  python scripts/generar_manual_despliegue.py
"""
import json
import os
import re
import sys

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "scripts"))
import modelo_objetivo
from modelo_objetivo import (MODELO, REGLAS, RETIRADAS, CLAVE_GENERADA,
                             CAMPOS_RETIRADOS, COLUMNAS_SIN_DECIDIR)
from inferencia import clasificar
from sistema import VOLCADO

# El modelo describe datos, no interfaz. Esto no se afirma de memoria: se
# comprueba. Si alguien declara vistas en el modelo, el manual deja de decir
# que no las hay en vez de quedarse mintiendo.
INTERFAZ = [n for n in ("VISTAS", "ACCIONES", "SLICES") if hasattr(modelo_objetivo, n)]

# Orden de construccion. Cada nivel solo referencia tablas de niveles anteriores
# o del suyo propio con el destino antes. Verificado topologicamente.
NIVELES = [
    ("1. Catalogos", ["ROL_Roles", "SED_Sedes", "UNF_UnidadesFuncionales", "CAL_Calzadas",
                      "SEN_Sentidos", "FRE_Frecuencias", "EST_Activo", "EOT_EstadosOrden",
                      "MOT_MotivosPendiente", "TPR_TiposRespuesta", "PAR_Parametros"]),
    ("2. Formularios", ["FRM_Formularios", "FRM_Secciones", "FRM_Preguntas", "LST_ValoresLista"]),
    ("3. Maestras", ["TIP_TiposActivo", "USR_Usuarios", "ASG_AsignacionZona", "FAL_ModosFalla"]),
    ("4. Activos", ["ACT_Activos"]),
    ("5. Ordenes", ["OT_OrdenesTrabajo", "PLA_PlanMantenimiento", "NOV_Novedades"]),
    ("6. Ejecucion", ["MAN_Mantenimientos", "CHK_Checklists", "CHD_ChecklistDetalle",
                      "FOT_Fotografias", "FIR_Firmas"]),
]

# ------------------------------------------------------------------- LAS CIFRAS
#
# Aqui vivia `TIPOS_MANUALES`: diecisiete columnas escritas a dedo bajo el
# titulo «los tipos que AppSheet no adivina». Era una LISTA BLANCA DE
# EXCEPCIONES sobre un default que se presumia bueno, y por eso el resto se
# daba por correcto por omision. Son 107 -las cuenta inferencia.py, no este
# archivo-, y dos de las diecisiete ni siquiera lo necesitaban: su propio
# nombre las consigue. El precio de la omision fue RG-03, bien escrita y bien
# colocada sobre una columna que AppSheet tipo Text.
#
# Regla de este generador: **ninguna cifra se escribe, todas se derivan.**

CLASES = clasificar()
TOTAL_COLUMNAS = sum(len(d["columnas"]) for d in MODELO.values())
REFS = [(t, c) for t in MODELO for c in MODELO[t]["columnas"] if c.get("ref")]
N_REFS = len(REFS)

REF_DE = {(t, c["nombre"]): c["ref"] for t, c in REFS}
CLAVE_DE = {c["nombre"]: t for t in MODELO
            for c in MODELO[t]["columnas"] if c.get("pk")}

# Referencias cuyo NOMBRE es ademas clave primaria de otra tabla. Son las que
# se prestan a abrirse en la tabla equivocada, y ahi AppSheet SI protesta:
# `contains a cyclical table reference`.
HOMONIMAS = [(t, c["nombre"], c["ref"], CLAVE_DE[c["nombre"]])
             for t, c in REFS if c["nombre"] in CLAVE_DE]


# --------------------------------------------- que reglas dependen de que columna
#
# Se resuelve por CADENA, no por nombre suelto. La diferencia no es cosmetica:
# `[Activo]` dentro del SELECT de RG-04 es ASG_AsignacionZona.Activo, no el
# `Activo` de las otras diecinueve tablas que se llaman igual. Contando por
# nombre salen 52 columnas con regla encima; contando por cadena, 23.
#
# No se importa de generar_prompt_expresiones.py porque ese modulo escribe su
# .md al importarse. Se reimplementa aqui, que es el coste de no tener efectos
# de importacion.
def _ambito(expresion):
    """Los tramos donde manda OTRA tabla: dentro de SELECT(Tabla[...], ...)."""
    tramos = []
    for m in re.finditer(r"\b(?:SELECT|FILTER|ANY|COUNT|MAXROW|MINROW)\s*\(\s*(\w+)\[",
                         expresion):
        tabla = m.group(1)
        if tabla not in MODELO:
            continue
        hondo, fin = 0, None
        for i in range(m.end() - len(tabla) - 1, len(expresion)):
            if expresion[i] == "(":
                hondo += 1
            elif expresion[i] == ")":
                hondo -= 1
                if hondo == 0:
                    fin = i
                    break
        tramos.append((m.start(), fin if fin is not None else len(expresion), tabla))
    return tramos


def _tocadas(tabla, expresion):
    """Los pares (tabla, columna) que la expresion realmente lee."""
    tramos = _ambito(expresion)
    out = set()
    for m in re.finditer(r"(?:\[\w+\]\s*\.\s*)*\[\w+\]", expresion):
        aqui = next((t for ini, fin, t in tramos if ini <= m.start() <= fin), tabla)
        for n in re.findall(r"\[(\w+)\]", m.group(0)):
            if aqui is None:
                break
            out.add((aqui, n))
            aqui = REF_DE.get((aqui, n))
    # SELECT nombra su columna fuera de corchetes, con la tabla delante.
    for m in re.finditer(r"\b(\w+)\[(\w+)\]", expresion):
        if m.group(1) in MODELO:
            out.add((m.group(1), m.group(2)))
    return out


DEPENDEN = {}
for _r in REGLAS:
    _pares = _tocadas(_r["tabla"], _r.get("expresion") or "")
    if _r.get("columna") and not _r["columna"].startswith("("):
        _pares.add((_r["tabla"], _r["columna"]))
    for _par in _pares:
        DEPENDEN.setdefault(_par, set()).add(_r["id"])


def reglas_de(tabla, columna):
    return sorted(DEPENDEN.get((tabla, columna), ()))


# ------------------------------------------------------- las que ESCRIBEN en la hoja
#
# Por TIPO de propiedad, no por lista de identificadores: una regla nueva entra
# sola en el grupo que le toca. Lo que escriben no se revierte cambiando un
# desplegable, y por eso van al final y con instantanea previa.
def escribe(r):
    return r["tipo"] in ("App formula", "Initial value") or r["tipo"].startswith("Bot")


ESCRIBEN = [r for r in REGLAS if escribe(r)]

# Orden de trabajo: primero lo que solo valida, al final lo que escribe.
REGLAS_ORDENADAS = sorted(REGLAS, key=lambda r: (escribe(r), r["id"]))


# ------------------------------------------------------------- lo que dice la hoja
#
# Se lee el volcado. Si no esta, la frase que dependa de ese dato NO se
# escribe: un manual que se inventa un numero cuando no puede leerlo es peor
# que un manual con un hueco.
_LIBRO = None


def _pestana(tabla):
    global _LIBRO
    if _LIBRO is None:
        try:
            import openpyxl
            _LIBRO = openpyxl.load_workbook(os.path.join(RAIZ, VOLCADO),
                                            read_only=True, data_only=True)
        except Exception:
            _LIBRO = False
    if not _LIBRO or tabla not in _LIBRO.sheetnames:
        return None
    filas = list(_LIBRO[tabla].iter_rows(values_only=True))
    if not filas:
        return None
    cab = [("" if x is None else str(x).strip()) for x in filas[0]]
    return [dict(zip(cab, f)) for f in filas[1:]]


_TIPOS = _pestana("TIP_TiposActivo")
RADIOS = {}
for _f in (_TIPOS or []):
    RADIOS.setdefault(_f.get("RadioGeofencingKm"), []).append(_f.get("Nombre"))
N_TIPOS = len(_TIPOS) if _TIPOS is not None else None
N_CON_RADIO = (sum(1 for _f in _TIPOS if _f.get("RadioGeofencingKm") not in (None, ""))
               if _TIPOS is not None else None)

_ACT = _pestana("ACT_Activos")
N_ACTIVOS = len(_ACT) if _ACT is not None else None


def _numero(x):
    """1.5 -> '1,5'. Sin decimales inventados."""
    return ("%g" % x).replace(".", ",")


# -------------------------------------------------- lo que dice la aplicacion viva
#
# La ultima instantanea guardada. NO es el destino: es el estado, y el manual
# tiene que distinguirlos porque confundirlos es de lo que mas cuesta aqui.
_CARPETA_FOTOS = os.path.join(RAIZ, "BD", "instantaneas")


def _ultima_foto():
    if not os.path.isdir(_CARPETA_FOTOS):
        return None, None
    fotos = [x for x in os.listdir(_CARPETA_FOTOS) if x.endswith(".json")]
    if not fotos:
        return None, None
    n = max(fotos, key=lambda x: os.path.getmtime(os.path.join(_CARPETA_FOTOS, x)))
    try:
        with open(os.path.join(_CARPETA_FOTOS, n), encoding="utf-8") as f:
            return n[:-5], json.load(f)
    except ValueError:
        return None, None


FOTO_NOMBRE, FOTO = _ultima_foto()
FOTO_FILAS = sum(len(v) for v in FOTO.values()) if FOTO else None
FOTO_VACIAS = sorted(t for t, v in FOTO.items() if not v) if FOTO else []

# Las columnas trampa NO se escriben a mano: se derivan mas abajo cruzando
# CAMPOS_RETIRADOS contra las claves del modelo. Una lista a mano aqui se
# desviaria del modelo en cuanto alguien lo edite.

L = []
w = L.append

w("# Manual de despliegue — SGMC sobre AppSheet")
w("")
w("**Para quien construye la aplicacion.** De cero a desplegada.")
w("")
w("> **Manual por rol, no por persona.** Quien lo ejecuta es el **Funcional**: perfil que configura")
w("> AppSheet, sin necesidad de programar. Escrito para poder replicarlo en otro contrato.")
w("")
w("| | |")
w("|---|---|")
w("| Sistema | Gestion de Mantenimiento en Campo |")
w("| Plataforma | Google AppSheet sobre Google Sheets |")
w("| Fuente del modelo | `scripts/modelo_objetivo.py`. Este manual se genera de ahi |")
w("| Tablas | **%d** |" % len(MODELO))
w("| Columnas | **%d** |" % TOTAL_COLUMNAS)
w("| Referencias | **%d** |" % N_REFS)
w("| Reglas | **%d** |" % len(REGLAS))
w("")
w("## Los cinco documentos, y cual se usa cuando")
w("")
w("**Este es el unico que hay que leer entero.** Los otros cuatro no son capitulos que falten")
w("aqui: son **extractos que se le pasan a un ejecutor** -una persona o un agente- para que haga")
w("una parte sin leer el resto. Si el que despliega es usted, le basta con este.")
w("")
w("| Documento | Que es | Cuando se usa | Describe |")
w("|---|---|---|---|")
w("| `MANUAL_DESPLIEGUE.md` | Este. La ruta completa, de cero a aplicacion funcionando | Siempre. Es el camino | el **destino** |")
w("| [`PROMPT_CABLEADO.md`](PROMPT_CABLEADO.md) | Encargo autocontenido de las %d referencias y de los tipos | Se copia integro a quien cablee | el **destino** |" % N_REFS)
w("| [`PROMPT_EXPRESIONES.md`](PROMPT_EXPRESIONES.md) | Idem para las %d reglas, con la **cadena de referencias** que atraviesa cada una | Se copia integro despues del cableado | el **destino** |" % len(REGLAS))
w("| [`TIPOS_ESPERADOS.md`](TIPOS_ESPERADOS.md) | La lista larga, tabla por tabla | Abierto al lado mientras se recorre *Data > Columns* | el **destino** |")
w("| [`CORRECCIONES_CABLEADO.md`](CORRECCIONES_CABLEADO.md) | **Generado contra la aplicacion viva** | Antes de empezar, y despues de cada tanda | el **estado de HOY** |")
w("")
w("**Cuatro describen el destino y uno el estado, y confundirlos es de lo que mas se paga.** Un")
w("documento generado del modelo seguira diciendo que hay %d referencias el dia que las %d esten" % (N_REFS, N_REFS))
w("puestas: no mira la aplicacion. **Solo `CORRECCIONES_CABLEADO.md` la mira.**")
w("")
w("## Por que este manual existe")
w("")
w("La primera version de esta aplicacion se construyo, y despues el modelo de datos se corrigio **en")
w("la hoja**: columnas renombradas, tablas nuevas, campos retirados. **No hubo forma de que AppSheet")
w("lo recogiera.**")
w("")
w("Dos limites de la plataforma, los dos verificados, explican por que:")
w("")
w("**`Regenerate` fusiona, no reemplaza.** Su documentacion dice que combina la informacion nueva con")
w("la existente e intenta mantener las columnas que ya estan. Sirve para anadir una columna; con un")
w("esquema muy divergente **impide converger**. El propio AppSheet indica la salida: *Delete and")
w("re-add the table*.")
w("")
w("**AppSheet ignora las pestanas ocultas, y no avisa.** Ocho pestanas del libro estaban ocultas y")
w("cargaban 24 tablas de 32, sin un solo mensaje.")
w("")
w("**Por debajo de cierto umbral se repara; por encima se reconstruye.** Este manual es el camino de")
w("reconstruir, que resulto ser mas rapido y mas limpio.")
w("")
w("---")
w("")
w("## Lo que este manual no puede saber")
w("")
w("**Todo lo que sigue sale del modelo, asi que describe el DESTINO: como tiene que quedar.** No")
w("sabe -ni puede saber- cuanto esta hecho ya. Si esta reconstruyendo de cero da igual; si esta")
w("retomando una aplicacion a medias, esa diferencia lo es todo.")
w("")
w("Al estado se le pregunta con dos comandos, y no estan aqui de adorno:")
w("")
w("```bash")
w("python scripts/auditar_cableado.py    # que referencias estan puestas HOY")
w("python scripts/instantanea.py         # que datos tiene la app HOY")
w("```")
w("")
w("**El auditor tiene un limite contraintuitivo, y conviene entenderlo antes de fiarse de el: es")
w("mas fuerte cuando el cableado esta MAL y mas debil cuando esta bien.** No lee el esquema -la")
w("API v2 devuelve filas, no esquema-: lee las columnas virtuales `Related ...` que AppSheet anade")
w("en la tabla DESTINO al crear una referencia. Si hay varias referencias entre el mismo par de")
w("tablas, AppSheet tiene que desambiguar y las llama `Related X By Columna`: ahi nombra la")
w("columna, y la lectura es prueba. Si hay una sola, solo nombra la tabla, y para saber que")
w("columna la produjo hay que preguntarselo al modelo, que es justo lo que se estaba verificando.")
w("")
w("> **De ahi la consecuencia practica: no sirve para confirmar una correccion recien hecha.** Si")
w("> alguien pone la referencia correcta en la columna equivocada, la virtual inversa se llama")
w("> igual y el auditor dice que esta bien. Lo explica entero el docstring de")
w("> `scripts/auditar_cableado.py`.")
w("")
w("**Y hay un tercer estado que no es ni bien ni mal: lo que no se puede ver.** La columna virtual")
w("vive en el destino, asi que de una referencia cuya tabla destino esta vacia el auditor no puede")
w("decir nada. No la da por buena: la separa.")
if FOTO:
    w("")
    w("La ultima instantanea guardada -`BD/instantaneas/%s.json`- trae **%d filas** repartidas en las"
      % (FOTO_NOMBRE, FOTO_FILAS))
    w("%d tablas, y **%d de ellas estan vacias**:" % (len(FOTO), len(FOTO_VACIAS)))
    w("")
    w("```")
    for _i in range(0, len(FOTO_VACIAS), 3):
        w("  " + " ".join("%-24s" % _t for _t in FOTO_VACIAS[_i:_i + 3]).rstrip())
    w("```")
    w("")
    w("**De esas %d salen dos cosas a la vez:** sus referencias no son medibles, y sus columnas se"
      % len(FOTO_VACIAS))
    w("tiparon a ciegas -sin contenido que leer, AppSheet cae en `Text`-. Las dos vuelven en el")
    w("paso 4.")
w("")
w("## Antes de nada: quien comprueba cada cosa")
w("")
w("**Quien ejecuta no puede verificarse a sí mismo.** Cierra el diálogo, ve el botón en gris, y")
w("para él la cosa quedó. El 2026-08-10 se reportó tres veces que algo estaba hecho:")
w("")
w("```")
w("«39/39 referencias asignadas»   ->  5 mal, 4 sin poner")
w("«11 reglas puestas»             ->  6 bien, 1 mal, 2 sin poner")
w("«tipos y Label listos»          ->  escribio en 2 celdas, 1 mal")
w("```")
w("")
w("Las tres se descubrieron después, una a una, y de ahí sale el bucle de arreglar el arreglo: cada")
w("tanda encuentra a mano lo que la anterior dio por hecho.")
w("")
w("Lo que las cazó no fue mirar más: fue **leer de vuelta con otro instrumento**.")
w("")
w("| Qué se toca | Quién lo lee de vuelta |")
w("|---|---|")
w("| Referencias | `python scripts/auditar_cableado.py` |")
w("| Datos | `python scripts/instantanea.py comparar <antes> <despues>` |")
w("| Estructura | `python scripts/verificar_app.py` |")
w("| **Tipos de columna** | **nadie: la API devuelve filas, no esquema** |")
w("| **Expresiones y filtros** | **nadie** |")
w("| **`Are updates allowed`** | **nadie**, y la API tiene más permisos que la app |")
w("| **`Label`** | **nadie** |")
w("")
w("### Y un orden que no es preferencia: los filtros de seguridad, al final")
w("")
w("**Poner un `Security Filter` apaga los instrumentos sobre esa tabla.** La API llama sin usuario,")
w("así que `USEREMAIL()` queda en blanco y el filtro no deja pasar nada:")
w("")
w("```")
w("ACT_Activos    368 filas  ->  0 por la API, en cuanto entra RG-04")
w("auditor        6 referencias no juzgables  ->  9")
w("```")
w("")
w("**No se pierde ni un dato** —un filtro filtra lecturas, no borra— pero `instantanea.py` deja de")
w("poder comparar los activos y `auditar_cableado.py` cuenta `ACT_Activos` como tabla vacía, con lo")
w("que las tres referencias que la apuntan dejan de ser juzgables.")
w("")
w("Así que `RG-04` y `RG-05` van **después** de haber comprobado referencias, tipos y datos. Es la")
w("versión instrumental de la trampa de siempre: no es que esté mal, es que **deja de poderse ver**,")
w("y eso se lee igual que «está bien» si nadie lo dice.")
w("")
w("> **Las cuatro de abajo son las que sobrevivieron a los tres informes.** No porque nadie mirara:")
w("> porque no había con qué. Se cierran copiando **literalmente** lo que muestra el editor, incluso")
w("> cuando coincide. «Coincide» no es evidencia; el texto sí.")
w("")

w("## Paso 0 — Antes de abrir AppSheet")
w("")
w("**Comprobar la hoja.** Si algo esta mal aqui, todo lo demas hereda el error.")
w("")
w("```bash")
w('python scripts/verificar_faseA.py "BD/<archivo>.xlsx"')
w("```")
w("")
w("Tiene que decir **`FASE A CERRADA`**. Si dice otra cosa, no siga.")
w("")
w("**Y mirar las pestanas ocultas**, que es lo que mas cuesta descubrir despues:")
w("")
w("```bash")
w('python -c "import openpyxl;wb=openpyxl.load_workbook(\'BD/<archivo>.xlsx\',read_only=True);'
  'print([n for n in wb.sheetnames if wb[n].sheet_state!=\'visible\'])"')
w("```")
w("")
w("Tiene que devolver **una lista vacia**. Si hay pestanas ocultas, mostrarlas en Google Sheets —")
w("*Ver → Hojas ocultas*— antes de continuar. **`F-18` de la verificacion tambien lo detecta.**")
w("")
w("## Paso 1 — Crear la aplicacion")
w("")
w("En AppSheet: **Create → App → Start with existing data**, y elegir el Google Sheets.")
w("")
w("**Fuente: el documento de Google Sheets, no un archivo subido.** Si se sube un `.xlsx`, la")
w("aplicacion queda leyendo una foto fija y nada se sincroniza.")
w("")
w("**Quien crea la aplicacion es su propietario.** Conviene que sea la cuenta que va a operarla: un")
w("coautor no puede dar de alta tablas, y todo este manual consiste en eso.")
w("")
w("## Paso 2 — Dar de alta las %d tablas" % len(MODELO))
w("")
w("*Data → `+` → Add data*, una por una. **En este orden**, que no es alfabetico: cada nivel apunta a")
w("tablas cuyas claves quedaron fijadas antes.")
w("")
for nombre, tablas in NIVELES:
    presentes = [t for t in tablas if t in MODELO]
    w("**%s**" % nombre)
    w("")
    w("```")
    for i in range(0, len(presentes), 3):
        w("  " + " · ".join("%-24s" % t for t in presentes[i:i+3]).rstrip())
    w("```")
    w("")
w("### Las %d que el modelo retira" % len(RETIRADAS))
w("")
w("**Sobre la hoja vigente estas pestanas ya no existen.** La hoja se genera del modelo, asi que no")
w("aparecen en el desplegable y no hay nada que evitar. La lista se conserva para reconocerlas si")
w("alguien trabaja sobre una copia antigua:")
w("")
w("| Pestana | Por que se retiro |")
w("|---|---|")
for t, motivo in sorted(RETIRADAS.items()):
    w("| `%s` | %s |" % (t, motivo))
w("")
w("**No lo de por hecho: compruebelo contra el archivo.** Tiene que devolver una lista vacia.")
w("")
w("```bash")
w('python -c "import openpyxl;n=openpyxl.load_workbook(\'BD/Modelo_Datos_PLANTILLA.xlsx\','
  "read_only=True).sheetnames;print([t for t in %s if t in n])\"" % sorted(RETIRADAS))
w("```")
w("")
w("**Y los bancos de preguntas que guardaban tres de ellas ya estan migrados** a `FRM_Preguntas`, que")
w("es el motor unico. Se comprueba contando cuantos formularios distintos tienen preguntas:")
w("")
w("```bash")
w('python -c "import openpyxl;s=openpyxl.load_workbook(\'BD/Modelo_Datos_PLANTILLA.xlsx\','
  "read_only=True,data_only=True)['FRM_Preguntas'];f=[r[1] for r in s.iter_rows(min_row=2,values_only=True)];"
  'print(len(f),\'preguntas en\',len(set(f)),\'formularios\')"')
w("```")
w("")
w("## Paso 3 — Las claves, todas `Text`")
w("")
w("*Data → Columns* de cada tabla. **Una sola casilla `KEY`**, sobre la columna correcta, tipo")
w("**`Text`**.")
w("")
w("| Tabla | Clave |")
w("|---|---|")
for t, d in sorted(MODELO.items()):
    pk = [c["nombre"] for c in d["columnas"] if c.get("pk")]
    if pk:
        w("| `%s` | `%s` |" % (t, pk[0]))
w("")
w("**`Text` sin excepcion, y hay un caso que lo justifica.** `USR_Usuarios.UsuarioID` tiene un valor")
w("alfanumerico entre otros numericos. Si AppSheet infiere `Number`, esa fila se queda sin clave")
w("valida y ese usuario **deja de existir para el sistema**.")
w("")
w("**Si ve dos casillas `KEY` marcadas, o la clave aparece como combinacion de dos columnas,")
w("corrijalo antes de seguir.** Contra una clave compuesta no resuelve ninguna referencia, y el")
w("sintoma es que falla todo el paso 5 sin decir por que.")
w("")
w("### Clave automatica para las filas nuevas")
w("")
w("Estas %d tablas crean filas desde la aplicacion. Sin esto, no sabe que identificador poner:" % len(CLAVE_GENERADA))
w("")
w("| Tabla | Columna | `Initial value` |")
w("|---|---|---|")
for t in sorted(CLAVE_GENERADA):
    pk = [c["nombre"] for c in MODELO[t]["columnas"] if c.get("pk")]
    if pk:
        w("| `%s` | `%s` | `UNIQUEID()` |" % (t, pk[0]))
w("")
w("## Paso 4 — El tipo de las %d columnas" % TOTAL_COLUMNAS)
w("")
w("**Subir el Excel arregla la hoja, no la aplicacion.** Son dos sitios distintos. El Excel fija")
w("que columnas hay y que datos tienen; **el tipo de cada columna vive en el esquema de")
w("AppSheet**, y ese se infiere. Reimportar no lo corrige: la inferencia vuelve a ser la misma")
w("sobre los mismos datos.")
w("")
w("### Este paso era una lista de excepciones, y por eso fallo")
w("")
w("Se titulaba «los tipos que AppSheet no adivina» y enumeraba unas pocas. Eso es una **lista")
w("blanca de excepciones sobre un default que se presume bueno**: lo que no salia en la lista se")
w("daba por correcto por omision, sin que nadie lo hubiera decidido. La plataforma garantiza lo")
w("contrario.")
w("")
w("**Lo que costo.** `RG-03` se puso bien escrita y bien colocada -`[CierreConExcepcion] = TRUE`")
w("en `Required_If`, sobre `MAN_Mantenimientos.MotivoExcepcion`- encima de una columna que")
w("AppSheet tipo `Text`. **Comparar texto contra el booleano `TRUE` es siempre falso y no da")
w("error:** el motivo de excepcion no se pide nunca, el tecnico cierra con excepcion sin")
w("justificar, y la regla figura como puesta. **Existe, esta bien redactada, y es decorativa.**")
w("")
w("Y el «que reportar» cerraba el bucle en falso -«cualquier tipo distinto del que dice este")
w("documento»-: **nadie puede reportar una diferencia contra un valor que nunca se le dio.**")
w("")
w("### Quien consigue el tipo de cada columna")
w("")
w("La pregunta util no es que tipos son raros, es **quien consigue este tipo**. Lo reparte")
w("`scripts/inferencia.py` y se comprueba en un comando:")
w("")
w("```bash")
w("python scripts/inferencia.py")
w("```")
w("")
w("| Quien lo consigue | Cuantas | Que significa |")
w("|---|---|---|")
w("| **Nadie: a mano** | **%d** | ningun contenido de la hoja las produce, o su propio nombre empuja a AppSheet al tipo equivocado |" % len(CLASES["a mano"]))
w("| El **nombre** | %d | la cabecera lleva una palabra que AppSheet reconoce |" % len(CLASES["nombre"]))
w("| El **contenido** | %d | AppSheet deberia acertar leyendo los valores, **cuando los hay** |" % len(CLASES["contenido"]))
w("")
w("**De donde sale la inferencia**, segun la documentacion oficial: AppSheet mira **el nombre de la")
w("cabecera Y el contenido de las filas**. Las palabras que disparan un tipo son concretas")
w("-`latlong` y `geolocation` para una coordenada, `birthday` o `day` para una fecha, una cabecera")
w("acabada en `?` para un Yes/No-. Ver `BASE_CONOCIMIENTO_APPSHEET.md` seccion 13.")
w("")
_con_regla = [x for x in CLASES["a mano"] if reglas_de(x[0], x[1]["nombre"])]
_por_tipo = {}
for _t, _c, _m in CLASES["a mano"]:
    _por_tipo[_c["tipo"]] = _por_tipo.get(_c["tipo"], 0) + 1
w("### Las %d que no consigue nadie si no las pone usted" % len(CLASES["a mano"]))
w("")
w("Por tipo: " + " · ".join("**%d** `%s`" % (n, t) for t, n in
                           sorted(_por_tipo.items(), key=lambda x: (-x[1], x[0]))) + ".")
w("")
w("**Estan ordenadas por cuantas reglas dependen de cada una**, que es lo que ordena el trabajo:")
w("una columna mal tipada **sin** regla encima molesta al usuario; **con** una regla encima")
w("**rompe la regla en silencio**, que es lo que paso con `RG-03`. Las %d primeras llevan regla."
  % len(_con_regla))
w("")
w("| Tabla | Columna | `TYPE` | Reglas | Por que no se consigue sola |")
w("|---|---|---|---|---|")
for _t, _c, _motivo in sorted(CLASES["a mano"],
                              key=lambda x: (-len(reglas_de(x[0], x[1]["nombre"])),
                                             x[0], x[1]["nombre"])):
    _det = ""
    if _c.get("ref"):
        _det = " → `%s`" % _c["ref"]
    elif _c.get("valores"):
        _det = " · " + " · ".join("`%s`" % v for v in _c["valores"])
    _rr = reglas_de(_t, _c["nombre"])
    w("| `%s` | `%s` | **`%s`**%s | %s | %s |"
      % (_t, _c["nombre"], _c["tipo"], _det,
         ", ".join("`%s`" % x for x in _rr) if _rr else "—", _motivo))
w("")
w("> **Ninguna de las %d referencias se creara sola, y es deliberado.** AppSheet infiere `Ref`" % N_REFS)
w("> cuando el nombre de una columna se parece al de una tabla, y las nuestras llevan prefijo:")
w("> `UNF_UnidadesFuncionales` no se parece a `UnidadFuncional`, asi que el parecido se rompe. Es")
w("> el precio de la convencion y a la vez su proteccion: impide que AppSheet invente referencias.")
w("> Como ponerlas es el paso 5.")
w("")
w("### Las %d que consigue el nombre" % len(CLASES["nombre"]))
w("")
w("Deberian entrar bien porque su cabecera lleva la palabra que AppSheet reconoce. **Compruebelas")
w("igual: es una heuristica, no una garantia.**")
w("")
for _t, _c, _motivo in sorted(CLASES["nombre"], key=lambda x: (x[0], x[1]["nombre"])):
    w("- `%s.%s` → **`%s`**  ·  %s" % (_t, _c["nombre"], _c["tipo"], _motivo))
w("")
w("### Las %d que dependen del contenido" % len(CLASES["contenido"]))
w("")
w("AppSheet deberia acertar leyendo los valores. La lista completa, tabla por tabla, esta en")
w("**[`TIPOS_ESPERADOS.md`](TIPOS_ESPERADOS.md)**, que es la que se tiene abierta al recorrer el")
w("editor. Aqui van las dos unicas cosas que hay que saber antes de mirarla:")
w("")
w("**Una columna vacia no tiene contenido que leer, y cae en `Text`.**")
if FOTO:
    w("Son %d tablas enteras -las de la instantanea- mas cada columna vacia de las pobladas."
      % len(FOTO_VACIAS))
w("")
w("> **Y el caso que desarma la confianza en el contenido.** Una columna de texto cuyos valores")
w("> parecen numeros se tipa `Number`. Paso el 2026-08-10 con `SED_Sedes.TramoINVIAS`: el unico")
w("> valor cargado era `5607`, asi que salio `Number` — y los otros tramos del corredor son")
w("> `55CN03`, que no cabe en un numero. **Tener el dato correcto no basta.**")
w("")
w("### Como se comprueba, tabla por tabla")
w("")
w("Al terminar de dar de alta o de regenerar una tabla, abrala en *Data > Columns* y **recorra la")
w("columna TYPE de arriba abajo contra la ficha del anexo**. No es opcional ni es paranoia: el")
w("defecto no se ve en la hoja ni en los datos, solo en esta pantalla. Lo que se corrige aqui")
w("sobrevive a un `Regenerate` posterior, porque AppSheet conserva el tipo de las columnas que ya")
w("existen.")
w("")
w("> **Si va a automatizarlo:** los desplegables de `TYPE` son `<select>` nativos del navegador, no")
w("> widgets propios, asi que se pueden asignar de forma determinista en vez de a base de clics.")
w("> **La senal de que la aplicacion recogio el cambio es que el boton `SAVE` pasa de gris a")
w("> azul**; si sigue gris, el cambio se perdera al recargar. Metodo y riesgos en")
w("> `BASE_CONOCIMIENTO_APPSHEET.md` seccion 15. No sustituye a comprobar: quita los clics, no la")
w("> verificacion.")
w("")
w("## Paso 5 — Las %d referencias" % N_REFS)
w("")
w("> **Cuidado con las listas de otros documentos.** Circulo una lista de **15** referencias por")
w("> convertir, y era correcta para lo que normaba: una aplicacion existente donde otras 23 ya estaban")
w("> puestas. Ese documento esta retirado. **Construyendo desde cero no sobrevive ninguna: son %d.**"
  % N_REFS)
w("> Si al terminar cuenta 15, siguio la lista equivocada.")
w("")
w("Una referencia de AppSheet **guarda el valor de la clave de la tabla destino**. De ahi que el orden")
w("importe: primero la clave del destino, despues quien la apunta.")
w("")
n = 0
for nombre, tablas in NIVELES:
    refs = [(t, c) for t in tablas if t in MODELO
            for c in MODELO[t]["columnas"] if c.get("ref")]
    if not refs:
        continue
    w("**%s**" % nombre)
    w("")
    w("```")
    for t, c in refs:
        n += 1
        marca = "   IsPartOf = TRUE" if c.get("es_parte_de") else ""
        w("%2d  %-34s -> %s%s" % (n, t + "." + c["nombre"], c["ref"], marca))
    w("```")
    w("")
w("**Nota sobre `OT_OrdenesTrabajo.OTOrigenID`**, que sale en el nivel 5: apunta a su propia tabla,")
w("para encadenar una orden derivada con la que la origino. **Dejela para el final del nivel.**")
w("")
w("### `IsPartOf` va marcado en cuatro, y en ninguna mas")
w("")
w("```")
for t, d in MODELO.items():
    for c in d["columnas"]:
        if c.get("es_parte_de"):
            w("%-34s -> %s" % (t + "." + c["nombre"], c["ref"]))
w("```")
w("")
w("**`MAN_Mantenimientos.OTID` va DESMARCADO, y es deliberado.** Con `IsPartOf`, borrar una orden")
w("borraria su ejecucion, sus fotografias y su firma **en cascada**. En un sistema cuyo proposito es")
w("que la evidencia sea dificil de falsificar, eso se decide, no se hereda de un ejemplo.")
w("")
w("### Despues de cada conversion")
w("")
w("**Mire si aparecieron celdas en blanco donde habia valores.** Convertir a `Ref` conserva solo las")
w("filas cuyo valor coincide con la clave del destino; las demas quedan huerfanas **sin mensaje de")
w("error**.")
w("")
w("### El error que AppSheet SI avisa: `cyclical table reference`")
w("")
w("`EstadoActivoID` es la **clave** de `EST_Activo` y a la vez el nombre de la **referencia** hacia")
w("ella en `ACT_Activos`. Son la misma palabra en dos tablas y no son la misma cosa. Si abre la")
w("columna en la tabla equivocada y la convierte alli, AppSheet lo rechaza:")
w("")
w("```")
w("Column Name 'EstadoActivoID' in Schema 'EST_Activo_Schema'")
w("contains a cyclical table reference to 'EST_Activo'.")
w("```")
w("")
w("Ya paso el 2026-08-10. **Y no es un caso raro: es el caso normal.** %d de las %d referencias"
  % (len(HOMONIMAS), N_REFS))
w("-el %d%%- llevan un nombre que ademas es clave primaria en otra tabla, porque es justo lo que"
  % round(100.0 * len(HOMONIMAS) / N_REFS))
w("produce la convencion `<Tabla>ID`.")
w("")
w("**Antes de tocar una columna, compruebe en que tabla esta.** De los tres fallos que persigue")
w("este manual, este es el unico que AppSheet le va a decir: el tipo mal inferido y la referencia")
w("bien puesta hacia el destino equivocado **no avisan**.")
w("")
w("## Paso 6 — RETIRADO. Sobre la hoja vigente no hay nada que deshacer")
w("")
w("> **No ejecute este paso.** Se conserva numerado para que quien tenga una copia antigua del manual")
w("> sepa que salio del plan, y para poder reconocer el problema si algun dia se trabaja sobre una")
w("> hoja heredada.")
w(">")
w("> Estas columnas **no existen en la hoja vigente**, que se genera del modelo, asi que AppSheet no")
w("> tiene nada que convertir solo. **Compruebelo usted, con la regla `F-19`:**")
w(">")
w("> ```bash")
w('> python scripts/verificar_faseA.py "BD/Modelo_Datos_PLANTILLA.xlsx"')
w("> ```")
w(">")
w("> ```")
w("> ok Hoja limpia: ninguna de las %d columnas retiradas existe ya. No hay nada que ocultar"
  % sum(len(v) for v in CAMPOS_RETIRADOS.values()))
w("> ```")
w(">")
w("> **Lo mismo vale para las marcas `OCULTAR` y `TRAMPA` del anexo:** describen una hoja que ya no se")
w("> usa y no aplican a la vigente.")
w("")
w("El problema que este paso resolvia: son columnas muertas que **se llaman igual que la clave de otra")
w("tabla**. Donde existan, AppSheet infiere la referencia por coincidencia de nombre y las convierte")
w("sin que nadie se lo pida.")
w("")
_cl = {}
for _t, _d in MODELO.items():
    for _c in _d["columnas"]:
        if _c.get("pk"):
            _cl[_c["nombre"]] = _t
_trampas = [(t, c, _cl[c], campos[c])
            for t, campos in sorted(CAMPOS_RETIRADOS.items())
            for c in sorted(campos) if c in _cl and _cl[c] != t]
w("| Tabla | Columna | Adonde apunta sola | Por que esta mal |")
w("|---|---|---|---|")
for t, c, destino, motivo in _trampas:
    w("| `%s` | `%s` | `%s` | %s |" % (t, c, destino, motivo))
w("")
w("**Son %d, derivadas del archivo y no escritas a mano.** Estan tambien en la ficha de cada tabla," % len(_trampas))
w("marcadas como TRAMPA, **y esas marcas tampoco aplican a la hoja vigente**.")
w("")
w("Si alguna vez aparecen —trabajando sobre una copia antigua del libro—, lo que habria que hacer es")
w("dejarlas en `Text` y desmarcar `Show?`. Como `Ref` dibujan rutas de navegacion que el modelo")
w("prohibe y aparecen en la aplicacion como si fueran buenas.")
w("")
w("## Paso 7 — Las %d reglas" % len(REGLAS))
w("")
w("Las expresiones enteras, con la **cadena de referencias que atraviesa cada una**, estan en")
w("[`PROMPT_EXPRESIONES.md`](PROMPT_EXPRESIONES.md) —que es lo que se le pasa a quien las ponga— y")
w("con su historia en [`sdd/RECONSTRUCCION_EXPRESIONES.md`](sdd/RECONSTRUCCION_EXPRESIONES.md) §2.")
w("**Copielas de ahi. No las escriba de memoria ni las adapte.**")
w("")
w("### Lo primero, porque ya salio mal tres veces")
w("")
w("**Una expresion con puntos no falla por estar mal escrita: falla porque un salto de su cadena")
w("no esta cableado.** El error de AppSheet lo dice literalmente. Cuando `RG-01` daba esto:")
w("")
w("```")
w("Can't find column \"RadioGeofencingKm\" in table \"SED_Sedes\"")
w("```")
w("")
w("...no habia que buscar otro nombre de columna: habia que ver **por que la cadena aterrizaba en")
w("`SED_Sedes`**. Y era que `ACT_Activos.TipoActivoID` apuntaba a la tabla de sedes.")
w("")
w("Cuando una falle: mire la tabla que nombra el error y busquela en la columna «atraviesa» de")
w("[`PROMPT_EXPRESIONES.md`](PROMPT_EXPRESIONES.md), que trae la cadena de cada regla desglosada")
w("salto a salto. **Ahi esta el salto roto**, y se arregla en el paso 5, no aqui.")
w("")
_radios = sorted((r for r in RADIOS if r not in (None, "")), reverse=True)
_detalle = " · ".join(
    "%s km (`%s`)" % (_numero(r), RADIOS[r][0]) if len(RADIOS[r]) == 1
    else "%s km en %d tipos" % (_numero(r), len(RADIOS[r]))
    for r in _radios)
w("> **NO reescriba una expresion para que el error desaparezca.** Se propusieron dos veces dos")
w("> arreglos que parecian razonables: sustituir el radio por un `LOOKUP` global a")
w("> `PAR_Parametros`, y quitar un salto de la cadena.")
if _radios:
    w("> **El primero habria colapsado en un solo numero los %d radios distintos** que hoy declara"
      % len(_radios))
    w("> `TIP_TiposActivo`: %s." % _detalle)
    w("> El segundo apunta a una columna que no existe.")
else:
    w("> El primero habria colapsado en un solo numero los radios distintos que declara")
    w("> `TIP_TiposActivo`, y el segundo apunta a una columna que no existe.")
w("> **Ninguno de los dos da error:** dejan el cierre en campo **aceptando lo que debe rechazar**.")
w("")
w("### Como se prueba una expresion")
w("")
w("En el **Asistente de Expresiones**, que solo evalua, y se cierra **sin dar a `Done`**. Escribir")
w("una expresion dentro de una columna para probarla la convierte en **configuracion activa**: ya")
w("ocurrio una vez, y dejo una `App formula` escribiendo coordenadas dentro de una columna")
w("retirada.")
w("")
w("**Y compruebe en que tabla esta antes de abrir la columna**, por lo que dice el paso 5: el")
w("mismo nombre es clave en una tabla y referencia en otra.")
w("")
w("### Las %d, y las %d que van al final" % (len(REGLAS), len(ESCRIBEN)))
w("")
w("| # | Regla | Tabla | Columna | Propiedad | Escribe |")
w("|---|---|---|---|---|---|")
for _i, _r in enumerate(REGLAS_ORDENADAS, 1):
    w("| %d | %s | `%s` | `%s` | %s | %s |"
      % (_i, _r.get("id", ""), _r.get("tabla", ""), _r.get("columna", "(tabla)"),
         _r.get("tipo", ""), "**SI**" if escribe(_r) else "no"))
w("")
w("### Las %d que escriben en la hoja van al final, y antes se toma una instantanea"
  % len(ESCRIBEN))
w("")
w("Las de `App formula`, `Initial value` y las de tipo bot **escriben en la hoja**. A diferencia de")
w("un tipo de columna, **lo que escriben no se revierte cambiando un desplegable**: hay que saber")
w("que habia antes. Por eso van las ultimas, cuando ya se puede comprobar que escribieron.")
w("")
w("```bash")
w("python scripts/instantanea.py guardar antes-de-las-que-escriben")
w("#   ... se ponen las %d ..." % len(ESCRIBEN))
w("python scripts/instantanea.py guardar despues")
w("python scripts/instantanea.py comparar antes-de-las-que-escriben despues")
w("```")
w("")
w("**Y no basta con mirar la fila que se espera que cambie.** Una `App formula` se evalua sobre")
_filas_act = len(FOTO.get("ACT_Activos", [])) if FOTO else N_ACTIVOS
if _filas_act:
    w("**todas** las filas de su tabla: `RG-16` sola se evalua sobre los **%d** activos, no sobre el"
      % _filas_act)
    w("unico que deberia cambiar. Si la expresion esta mal, escribe en todos y **no da error: da")
    w("datos**. Por eso el criterio de cierre no es «la fila esperada quedo bien», es **«no cambio")
    w("ninguna celda»** — y eso exige la fotografia previa.")
else:
    w("**todas** las filas de su tabla, no solo sobre la que se espera que cambie. Si la expresion")
    w("esta mal, escribe en todas y **no da error: da datos**.")
w("")
w("### Las cuatro que no pueden faltar")
w("")
w("**Geofencing** — en `MAN_Mantenimientos.Coordenadas_Cierre_LatLong`:")
w("")
w("```")
w("Initial value:  HERE()")
w("Valid_If:       DISTANCE([Coordenadas_Cierre_LatLong], [OTID].[ActivoID].[Ubicacion_LatLong])")
w("                  <= [OTID].[ActivoID].[TipoActivoID].[RadioGeofencingKm]")
w("Invalid text:   Ubicacion fuera de rango: debe estar junto al activo para cerrar.")
w("Editable_If:    FALSE")
w("```")
w("")
w("**El radio va por tipo de activo, no como literal.** Una subestacion y un poste SOS no admiten")
w("la misma tolerancia, y un tramo de fibra es lineal. `PAR_Parametros.RADIO_GEOFENCING_KM` queda")
w("como valor provisional historico: **la regla no lo lee.**")
w("")
w("**Antes de pegarla, compruebe que la columna esta poblada**, porque contra celdas en blanco esta")
w("expresion **rechaza tambien los cierres legitimos**:")
w("")
w("```bash")
w('python -c "import openpyxl;s=openpyxl.load_workbook(\'%s\','
  "read_only=True,data_only=True)['TIP_TiposActivo'];h=[c.value for c in next(s.iter_rows(max_row=1))];"
  "i=h.index('RadioGeofencingKm');v=[r[i] for r in s.iter_rows(min_row=2,values_only=True)];"
  'print(len(v),\'tipos,\',sum(1 for x in v if x not in (None,\'\')),\'con radio\')"' % VOLCADO)
w("```")
w("")
if N_TIPOS is not None:
    w("Sobre la hoja vigente devuelve **%d tipos, %d con radio**. Si devuelve alguno sin radio, pare:"
      % (N_TIPOS, N_CON_RADIO))
    w("ese tipo de activo no se podra cerrar en campo.")
else:
    w("Los dos numeros tienen que coincidir. Si devuelve alguno sin radio, pare: ese tipo de activo")
    w("no se podra cerrar en campo.")
w("")
w("**No editables** — en `MAN_Mantenimientos`, `Editable_If = FALSE` en las cuatro columnas de")
w("captura:")
w("")
w("```")
w("Coordenadas_Cierre_LatLong · Precision_GPS · UbicacionEscaneo_LatLong · FechaHoraEscaneo")
w("```")
w("")
w("**Sin esto el geofencing es decorativo:** el tecnico arrastra el pin del mapa y cierra desde")
w("donde quiera. La regla parece funcionar y no prueba nada.")
w("")
w("> **Supuesto sin verificar, y es el peor modo de fallo del sistema.** No hay pagina oficial que")
w("> confirme si AppSheet evalua un `Valid_If` sobre una columna con `Editable_If = FALSE`. **Si no")
w("> lo evalua, la regla parece funcionar por no ejercitarse nunca.** Se detecta asi: pruebe un")
w("> cierre cercano y uno lejano. **Si los dos salen aceptados, sospeche de esto antes que del")
w("> radio.**")
w("")
w("**Excepcion por GPS deficiente** — en `MAN_Mantenimientos.CierreConExcepcion`:")
w("")
w("```")
w('App formula:')
w('OR(ISBLANK(LOOKUP("UMBRAL_GPS", "PAR_Parametros", "ParametroID", "Valor")),')
w('   [Precision_GPS] > LOOKUP("UMBRAL_GPS", "PAR_Parametros", "ParametroID", "Valor"))')
w("```")
w("")
w("**El `ISBLANK` no sobra.** Sin el, borrar la fila del parametro hace que **todos los cierres")
w("salgan limpios y nadie se entere**. Con el, si el umbral no se puede leer el cierre se marca")
w("como excepcional: falla hacia el lado seguro.")
w("")
w("**Y esta columna es la del paso 4.** Si `CierreConExcepcion` quedo `Text`, esta `App formula` le")
w("escribe texto y `RG-03` —que compara contra el booleano `TRUE`— deja de pedir el motivo. Dos")
w("reglas puestas, ninguna de las dos haciendo nada, y ni un mensaje de error.")
w("")
w("**Filtros de seguridad** — *Data → Tables → [tabla] → Security Filter*:")
w("")
w("```")
w("ACT_Activos:")
w("IN([UnidadFuncionalID], SELECT(ASG_AsignacionZona[UnidadFuncionalID],")
w("   AND([UsuarioID].[Correo] = USEREMAIL(), [Activo] = TRUE)))")
w("")
w("OT_OrdenesTrabajo:")
w("OR([TecnicoID].[Correo] = USEREMAIL(), [SupervisorID].[Correo] = USEREMAIL())")
w("```")
w("")
w("**No son solo control de acceso: son rendimiento.** Sin ellos, cada tecnico se descarga el")
w("inventario entero al telefono.")
w("")
w("### Retirar el borrado — sin esto el `IsPartOf` es peligroso")
w("")
w("*Data → Tables → [tabla] → Are updates allowed*:")
w("")
w("```")
w("OT_OrdenesTrabajo    Updates si · Adds si · Deletes NO")
w("MAN_Mantenimientos   Updates si · Adds si · Deletes NO")
w("```")
w("")
w("**Es la otra mitad del paso 5.** Marcar `IsPartOf` en cuatro referencias crea **borrado en")
w("cascada**: borrar un mantenimiento se lleva sus fotografias, su firma y su checklist.")
w("")
w("Eso solo es seguro **porque el mantenimiento nunca se borra**, y eso es exactamente lo que hace")
w("quitar `Deletes`. Configurar el `IsPartOf` sin esto deja la cascada abierta.")
w("")
w("## Paso 8 — Las vistas")
w("")
w("**Este manual no especifica las vistas, y hay que saberlo antes de empezar el paso.** El modelo")
if INTERFAZ:
    w("declara ahora %s en `scripts/modelo_objetivo.py`: **actualice este generador**, porque el paso"
      % " y ".join("`%s`" % n for n in INTERFAZ))
    w("sigue escrito como si no existieran.")
else:
    w("declara datos, no interfaz: `VISTAS`, `ACCIONES` y `SLICES` **no existen** en")
    w("`scripts/modelo_objetivo.py` —comprobado al generar este manual, no de memoria—. Por eso aqui no")
    w("hay ficha columna por columna como en los pasos anteriores, y por eso lo que decida en este paso")
    w("es lo unico que no queda escrito en ninguna parte.")
w("")
w("**Lo unico que AppSheet crea solo son las columnas virtuales `Related ...`**, que aparecen al poner")
w("las referencias del paso 5 y traen con ellas la navegacion padre-hijo: al abrir un mantenimiento se")
w("ven sus fotografias y su firma; al abrir un activo, sus ordenes. **Eso es todo lo que se construye")
w("solo.** Las pantallas no.")
w("")
w("**Y no se configuran a ojo.** Las tres de abajo van con el tipo y la tabla que dice la ficha; si")
w("una no encaja, no la improvise: **anote que falta y siga**. Una vista inventada aqui es")
w("configuracion activa que nadie declaro y que el siguiente que reconstruya la aplicacion no podra")
w("reproducir.")
w("")
w("| Vista | Tipo | Sobre | Nota |")
w("|---|---|---|---|")
w("| Mapa de activos | `Map` | `ACT_Activos` | Columna de mapa: `Ubicacion_LatLong` |")
w("| Mis ordenes | `Deck` | `OT_OrdenesTrabajo` | Es la pantalla de trabajo del tecnico |")
w("| Mantenimientos | `Table` | `MAN_Mantenimientos` | |")
w("")
w("**Anote lo que haga.** Es la unica constancia que va a quedar de este paso.")
w("")
w("## Paso 9 — Verificar antes de publicar")
w("")
w("**No lo de por cerrado usted.** Este proyecto tiene tres cierres reportados que no resistieron la")
w("comprobacion contra el archivo, y las tres veces lo paro un script.")
w("")
w("**La cadena navega** — en el Asistente de Expresiones, sobre `MAN_Mantenimientos`:")
w("")
w("```")
w("[OTID].[ActivoID].[Ubicacion_LatLong]")
w("[OTID].[TecnicoID].[Correo]")
w("```")
w("")
w("Las dos en verde. Si la primera falla, casi siempre es `OT_OrdenesTrabajo.ActivoID`, que la lista")
w("antigua de 15 no incluia.")
w("")
w("**Cuente las referencias.** Las columnas de tipo `Ref` deben sumar **%d**."
  % N_REFS)
w("")
w("**Y los seis verificadores del repositorio:**")
w("")
w("```bash")
w("python scripts/validar_modelo.py          # el modelo consigo mismo")
w('python scripts/verificar_faseA.py "..."   # el modelo contra la hoja')
w("python scripts/verificar_documentos.py    # la prosa contra el modelo")
w("python scripts/verificar_enlaces.py       # que todo enlace entre documentos resuelve")
w("python scripts/verificar_reproducible.py  # que generar dos veces da el mismo archivo")
w("python scripts/verificar_datos.py         # que los DATOS sostienen lo que el modelo declara")
w("```")
w("")
w("**Ninguno de los seis mira la aplicacion:** todos comparan el modelo, la hoja y la prosa entre")
w("si. Los dos que si la miran son los del principio —`auditar_cableado.py` e `instantanea.py`—, y")
w("ninguno de los dos lee el esquema, porque la API devuelve filas. **De lo que hay configurado en")
w("el editor, la unica prueba es haberlo mirado ahi.** Para lo demas estan las pruebas de")
w("aceptacion de [`sdd/PRUEBA-003-despliegue.md`](sdd/PRUEBA-003-despliegue.md).")
w("")
w("## Paso 10 — Publicar")
w("")
w("> **Antes de publicar, lea esto.** Ninguna de las coordenadas de `ACT_Activos` se levanto en campo.")
_moda = max((r for r in RADIOS if r not in (None, "")), key=lambda r: len(RADIOS[r]),
            default=None)
w("> Las **%s** se **derivan del PK sobre el trazado del corredor** en cada pasada del generador:"
  % (N_ACTIVOS if N_ACTIVOS is not None else "de ACT_Activos"))
w("> son todas distintas y todas estan sobre la via, pero **ninguna esta medida**. Con los radios")
if _moda is not None:
    w("> por tipo —%s km en %d de los %d tipos— la aplicacion"
      % (_numero(_moda), len(RADIOS[_moda]), N_TIPOS))
else:
    w("> por tipo, tan pequenos como son, la aplicacion")
w("> **rechaza todo cierre hecho en via y acepta todo cierre hecho en Bogota**.")
w(">")
w("> **No es un defecto de la configuracion: faltan las coordenadas reales**, que es la decision")
w("> D-01. Publicar antes de cargarlas entrega un sistema donde ningun tecnico puede cerrar una")
w("> orden, y se descubre con el tecnico delante.")
w("")
w("*Manage → Deploy → Run deployment check*, y despues **Move app to Deployed state**.")
w("")
w("**Antes de publicar, si existe una aplicacion anterior sobre la misma hoja, despubliquela.** Dos")
w("aplicaciones sobre un backend sin integridad referencial es una fuente de corrupcion silenciosa:")
w("la vieja conserva permisos de anadir y borrar que el modelo nuevo ya no concede.")
w("")
w("## Reversion — hasta donde se puede volver atras")
w("")
w("**Todo lo anterior al paso 10 se puede abandonar sin coste.** La aplicacion no esta publicada y")
w("nadie la usa: se borra y se empieza de nuevo. La hoja no se toca en ningun paso salvo el 0.")
w("")
w("**El paso 0 SI escribe en la hoja** al mostrar las pestanas ocultas. Antes de empezar, haga una")
w("copia fechada del documento. Es el unico punto de restauracion del dato.")
w("")
w("**El punto de no retorno es el paso 10**, y no por publicar: por **despublicar la aplicacion")
w("anterior**. Si el *deployment check* falla despues, la vieja ya no esta en servicio. Compruebe")
w("todo el paso 9 **antes** de despublicar nada.")
w("")
w("### Lo que no se puede deshacer con un comando")
w("")
w("**Antes de cambiar un tipo o una referencia, anote el valor que tenia.** Suena a burocracia y")
w("no lo es: **no hay ningun comando en este repositorio que lo recupere despues.** La API v2 de")
w("AppSheet devuelve **filas, no esquema**, asi que no se le puede preguntar de que tipo era una")
w("columna ni adonde apuntaba una referencia. Es el mismo limite por el que")
w("`auditar_cableado.py` tiene que leer de rebote las columnas virtuales `Related ...`.")
w("")
w("Una linea por cambio basta, y es lo unico que hay:")
w("")
w("```")
w("ACT_Activos.TipoActivoID    estaba: Ref -> SED_Sedes     lo dejo: Ref -> TIP_TiposActivo")
w("SED_Sedes.TramoINVIAS       estaba: Number               lo dejo: Text")
w("```")
w("")
w("**Y para las %d reglas que escriben en la hoja, la anotacion no sirve: hace falta la"
  % len(ESCRIBEN))
w("instantanea.** Lo que escriben no vive en el esquema, vive en el dato, y hay que haberlo")
w("fotografiado **antes**:")
w("")
w("```bash")
w("python scripts/instantanea.py guardar antes-de-las-que-escriben")
w("```")
w("")
w("Sin esa foto, la comparacion posterior no tiene contra que compararse, y una `App formula`")
w("equivocada no deja rastro de error: deja datos.")
w("")
w("## Lo que NO cabe en el plan gratuito")
w("")
w("No es *mas adelante*: es **no en este plan**. Solo cambia con la decision de licenciamiento.")
w("")
w("| Lo que se querria | Por que no |")
w("|---|---|")
w("| Generacion automatica de las ordenes del mes | Los procesos programados no se ejecutan |")
w("| Aviso al supervisor de que hay trabajo por recibir | Lo mismo |")
w("| Integracion con sistemas externos | Sin plan Core no hay API REST |")
w("| Atributos distintos por tipo de equipo | El backend es una hoja: no hay esquema dinamico |")
w("| Que una escritura directa en la hoja respete las validaciones | Imposible por diseno |")
w("")
w("**Ese ultimo importa mas de lo que parece.** Todas las garantias del sistema viven en la capa de")
w("aplicacion. Quien escriba en la hoja se las salta todas. Lo que el sistema puede ofrecer es que")
w("falsificar cueste mas que hacer el trabajo, no que sea imposible.")
w("")
w("---")
w("*Generado de `scripts/modelo_objetivo.py` por `scripts/generar_manual_despliegue.py`.*")
w("*Para actualizarlo, cambie el modelo y vuelva a generar.*")


# ------------------------------------------------------------------ ANEXO
_claves = {}
for _t, _d in MODELO.items():
    for _c in _d["columnas"]:
        if _c.get("pk"):
            _claves[_c["nombre"]] = _t

w("---")
w("")
w("# Anexo — Ficha de cada tabla")
w("")
w("**Columna por columna, sin nada que deducir.** Esta es la referencia contra la que se configura y")
w("contra la que se valida. Si una columna no aparece aqui, no deberia estar visible en la app.")
w("")
w("> ## Las marcas `OCULTAR` y `TRAMPA` NO aplican a la hoja vigente")
w(">")
w("> **Describen una hoja que ya no se usa:** el libro heredado que arrastraba columnas que el modelo")
w("> no declara. **La hoja vigente se genera del modelo y no trae ninguna**, asi que no hay nada que")
w("> ocultar ni ninguna referencia que deshacer. Ignore las dos marcas.")
w(">")
w("> Se conservan por una sola razon: son la lista por nombre que permite reconocer esas columnas si")
w("> algun dia aparece una copia antigua del libro. **No son trabajo de nadie.**")
w(">")
w("> **Compruebelo, con la regla `F-19`:**")
w(">")
w("> ```bash")
w('> python scripts/verificar_faseA.py "BD/Modelo_Datos_PLANTILLA.xlsx"')
w("> ```")
w(">")
w("> ```")
w("> ok Hoja limpia: ninguna de las %d columnas retiradas existe ya. No hay nada que ocultar"
  % sum(len(v) for v in CAMPOS_RETIRADOS.values()))
w("> ```")
w(">")
w("> La plantilla se rehace entera con `python scripts/generar_plantilla.py` y sale con las columnas")
w("> que el modelo declara, ni una mas.")
w("")
w("Leyenda:")
w("")
w("- **CLAVE** — casilla `KEY` marcada, tipo `Text`")
w("- **`Ref` -> Tabla** — tipo `Ref` con esa tabla como *Source table*")
w("- **IsPartOf** — ademas, casilla `Is a part of` marcada")
w("- **OCULTAR** — **no aplica a la hoja vigente.** Columna retirada del modelo que el libro heredado")
w("  arrastraba. Si apareciera: tipo `Text`, `Show?` desmarcado, sin formula")
w("- **TRAMPA** — **no aplica a la hoja vigente.** Donde exista, AppSheet la convierte a `Ref` sola")
w("  por coincidencia de nombre")
w("- **SIN DECIDIR** — esta en la hoja y el modelo no la declara")
w("")

for tabla in sorted(MODELO):
    d = MODELO[tabla]
    w("## `%s`" % tabla)
    w("")
    w(d["proposito"])
    w("")
    w("| Columna | Tipo | Que hacer |")
    w("|---|---|---|")
    for c in d["columnas"]:
        acc = []
        if c.get("pk"):
            acc.append("**CLAVE**")
        if c.get("ref"):
            acc.append("`Ref` -> `%s`" % c["ref"])
            acc.append("**IsPartOf**" if c.get("es_parte_de") else "IsPartOf desmarcado")
        # Los valores del Enum salen de `valores`, NUNCA de la nota. Partir la
        # nota por comas publicaba un valor llamado "Baja. Pondera la
        # disponibilidad de D-13" y otro con un parrafo entero dentro.
        if c["tipo"] == "Enum":
            if c.get("valores"):
                acc.append("Valores: " + " · ".join("`%s`" % v for v in c["valores"]))
            else:
                acc.append("**Valores sin declarar en el modelo.** No los invente: "
                           "pregunte antes de crear la columna")
        if c.get("valor_inicial"):
            acc.append("`Initial value` = `%s`" % c["valor_inicial"])
        w("| `%s` | `%s` | %s |" % (c["nombre"], c["tipo"], " · ".join(acc)))
    ret = CAMPOS_RETIRADOS.get(tabla, {})
    sind = [c for (t2, c) in COLUMNAS_SIN_DECIDIR if t2 == tabla]
    if ret or sind:
        w("")
        w("**Y estas, que estan en la hoja y NO se usan:**")
        w("")
        w("| Columna | Que hacer | Por que |")
        w("|---|---|---|")
        for c, motivo in sorted(ret.items()):
            tr = " · **TRAMPA: AppSheet la pone `Ref` sola hacia `%s`**" % _claves[c] if c in _claves else ""
            w("| `%s` | **OCULTAR**%s | %s |" % (c, tr, motivo))
        for c in sorted(sind):
            w("| `%s` | **OCULTAR** · SIN DECIDIR | El modelo no la declara |" % c)
    w("")

salida = os.path.join(RAIZ, "docs", "MANUAL_DESPLIEGUE.md")
with open(salida, "w", encoding="utf-8") as f:
    f.write("\n".join(L) + "\n")

print("Generado:", salida)
print("%d tablas · %d columnas · %d referencias · %d reglas · %d claves generadas"
      % (len(MODELO), TOTAL_COLUMNAS, N_REFS, len(REGLAS), len(CLAVE_GENERADA)))
print("Tipos: %d a mano · %d por nombre · %d por contenido   (inferencia.py)"
      % (len(CLASES["a mano"]), len(CLASES["nombre"]), len(CLASES["contenido"])))
print("%d de las %d a mano llevan al menos una regla encima"
      % (len([x for x in CLASES["a mano"] if reglas_de(x[0], x[1]["nombre"])]),
         len(CLASES["a mano"])))
print("%d de las %d reglas escriben en la hoja · %d de las %d referencias llevan nombre "
      "de clave ajena" % (len(ESCRIBEN), len(REGLAS), len(HOMONIMAS), N_REFS))
if FOTO:
    print("Estado leido de BD/instantaneas/%s.json: %d filas, %d tablas vacias"
          % (FOTO_NOMBRE, FOTO_FILAS, len(FOTO_VACIAS)))
else:
    print("Sin instantanea en BD/instantaneas/: el manual sale sin las cifras de estado")
