# -*- coding: utf-8 -*-
"""Verifica la Fase A contra un .xlsx exportado del Sheets de produccion.

Compara encabezado por encabezado contra el modelo objetivo. No compara contra
lo que alguien reporto: compara contra scripts/modelo_objetivo.py.

Uso:  python scripts/verificar_faseA.py "BD/Modelo de Datos (4).xlsx"
Salida: 0 si la Fase A esta cerrada, 1 si falta algo.

Por que existe
--------------
La Fase A la aplico un asistente distinto, a mano sobre la hoja, y se reporto
como cerrada al 100%. Este proyecto arrastra un historial de subsanaciones
reportadas como cerradas que no lo estaban, de modo que el reporte no cierra
nada: lo cierra el archivo.
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from modelo_objetivo import (MODELO, RENOMBRADOS, RETIPADOS, CAMPOS_RETIRADOS,
                             CLAVE_LEGIBLE, CLAVE_GENERADA, PARAMETROS)

try:
    import openpyxl
except ImportError:
    print("Falta openpyxl."); sys.exit(2)

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# Por defecto, la hoja que la aplicacion lee de verdad. Estuvo apuntando a
# "Modelo de Datos (4).xlsx" mucho despues de que ese libro dejara de ser el
# vigente: correrlo sin argumento daba un veredicto sobre un archivo muerto.
from sistema import VOLCADO
ruta = sys.argv[1] if len(sys.argv) > 1 else os.path.join(RAIZ, VOLCADO)
if not os.path.isabs(ruta):
    ruta = os.path.join(RAIZ, ruta)

# DOS libros a proposito. Sin data_only, openpyxl devuelve el TEXTO de la
# formula, no su valor: TIP_TiposActivo.FormularioID daba 18 huerfanos contra
# FRM_Formularios y ninguna regla lo veia. Con data_only, en cambio, la formula
# desaparece y F-17 no tendria nada que detectar. Hacen falta los dos.
wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)    # valores
wb_formulas = openpyxl.load_workbook(ruta, read_only=True)           # formulas

fallos, avisos, oks = [], [], []


def falla(codigo, msg):
    fallos.append("[%s] %s" % (codigo, msg))


def aviso(codigo, msg):
    avisos.append("[%s] %s" % (codigo, msg))


def encabezados(hoja):
    ws = wb[hoja]
    fila = next(ws.iter_rows(min_row=1, max_row=1))
    return [str(c.value).strip() for c in fila if c.value is not None]


def filas(hoja):
    ws = wb[hoja]
    return sum(1 for r in ws.iter_rows(min_row=2, values_only=True)
               if any(v not in (None, "") for v in r))


# ------------------------------------------------ F-01 los renombrados llegaron
for tabla, mapa in RENOMBRADOS.items():
    if tabla not in wb.sheetnames:
        falla("F-01", "%s no existe en el libro" % tabla)
        continue
    h = encabezados(tabla)
    for viejo, (nuevo, _m) in mapa.items():
        if nuevo not in h:
            falla("F-01", "%s.%s no existe. El renombrado desde '%s' no se aplico"
                  % (tabla, nuevo, viejo))
            continue
        oks.append("%s.%s" % (tabla, nuevo))
        # El nombre viejo solo puede sobrevivir si el modelo objetivo lo reutiliza
        # para otra cosa. Es el caso de OT_OrdenesTrabajo.Activo, que pasa de ser
        # el vinculo al activo a ser la bandera Yes/No.
        if viejo in h:
            reutilizado = any(c["nombre"] == viejo for c in MODELO[tabla]["columnas"])
            if reutilizado:
                aviso("F-01", "%s.%s sigue existiendo, pero el modelo lo reutiliza como "
                              "columna propia. Correcto, no es un fallo" % (tabla, viejo))
            else:
                falla("F-01", "%s.%s sigue con el nombre viejo" % (tabla, viejo))

# --------------------------------------- F-02 toda columna del modelo esta o no
RETIRADAS_OK = {t: set(c) for t, c in CAMPOS_RETIRADOS.items()}
for tabla, d in MODELO.items():
    if tabla not in wb.sheetnames:
        falla("F-02", "%s no existe en el libro" % tabla)
        continue
    h = set(encabezados(tabla))
    faltan = [c["nombre"] for c in d["columnas"] if c["nombre"] not in h]
    if faltan:
        falla("F-02", "%s: faltan %d columnas del modelo: %s"
              % (tabla, len(faltan), ", ".join(faltan)))

# ------------------------------------ F-03 lo retirado sigue ahi, y esta bien
for tabla, campos in CAMPOS_RETIRADOS.items():
    if tabla not in wb.sheetnames:
        continue
    h = set(encabezados(tabla))
    vivos = [c for c in campos if c in h]
    if vivos:
        aviso("F-03", "%s conserva %d columnas marcadas como retiradas: %s. "
                      "Es lo esperado: la Fase A no borra nada"
              % (tabla, len(vivos), ", ".join(vivos)))

# --------------------------------------------- F-04 lo retipado sigue pendiente
n_retipar = sum(len(m) for m in RETIPADOS.values())
aviso("F-04", "%d columnas siguen pendientes de retipar a Ref. Es trabajo de la "
              "Fase B, en el editor de AppSheet, no de la hoja" % n_retipar)

# ------------------------------------------------------ F-05 limpieza del CHK
if "CHK_Checklists" in wb.sheetnames:
    ws = wb["CHK_Checklists"]
    ids = [str(r[0]).strip() for r in ws.iter_rows(min_row=2, values_only=True)
           if r and r[0] not in (None, "")]
    # d02d8a3d se retira de la lista de conservadas y pasa a la de borradas: ESPEC-001C
    # ordeno eliminarla porque su MantenimientoID guardaba 'OT-0001', que es una ORDEN.
    # No basta con dejar de exigir que exista; hay que exigir que NO exista.
    for basura in ("CHK001", "0356e6d7", "d02d8a3d"):
        if basura in ids:
            falla("F-05", "CHK_Checklists: '%s' sigue ahi. Debia borrarse" % basura)
    if not any(b in ids for b in ("CHK001", "0356e6d7", "d02d8a3d")):
        oks.append("CHK_Checklists: las tres filas de ensayo estan borradas")

# ---------------------------------------------- F-06 catalogos nuevos poblados
POBLACION_MINIMA = {
    "UNF_UnidadesFuncionales": (4, "las cuatro unidades funcionales, con claves 7 a 10"),
    "EOT_EstadosOrden": (7, "los siete estados del ciclo de vida"),
    "MOT_MotivosPendiente": (5, "los cinco motivos del supuesto D-07"),
    "ASG_AsignacionZona": (1, "al menos una asignacion, o el Security Filter deja a "
                              "cada tecnico en cero activos"),
}
for tabla, (minimo, para_que) in POBLACION_MINIMA.items():
    if tabla not in wb.sheetnames:
        falla("F-06", "%s no existe" % tabla)
        continue
    n = filas(tabla)
    if n < minimo:
        falla("F-06", "%s tiene %d filas y necesita al menos %d: %s"
              % (tabla, n, minimo, para_que))
    else:
        oks.append("%s poblada con %d filas" % (tabla, n))

# -------------------------------------- F-07 las claves de UNF son las de ACT
if "UNF_UnidadesFuncionales" in wb.sheetnames and "ACT_Activos" in wb.sheetnames:
    ws = wb["UNF_UnidadesFuncionales"]
    claves = {str(r[0]).strip() for r in ws.iter_rows(min_row=2, values_only=True)
              if r and r[0] not in (None, "")}
    h = encabezados("ACT_Activos")
    if "UnidadFuncionalID" in h:
        i = h.index("UnidadFuncionalID")
        ws2 = wb["ACT_Activos"]
        usados = {str(r[i]).strip() for r in ws2.iter_rows(min_row=2, values_only=True)
                  if r and len(r) > i and r[i] not in (None, "")}
        huerfanos = usados - claves
        if huerfanos:
            falla("F-07", "ACT_Activos.UnidadFuncionalID usa %s, que no existe en "
                          "UNF_UnidadesFuncionales. La conversion a Ref dejaria esas "
                          "filas huerfanas" % sorted(huerfanos))
        else:
            oks.append("Las %d unidades funcionales distintas usadas por ACT_Activos resuelven "
                       "contra UNF" % len(usados))

# ------------------------------- F-08 los estados de la orden resuelven contra EOT
if "EOT_EstadosOrden" in wb.sheetnames and "OT_OrdenesTrabajo" in wb.sheetnames:
    ws = wb["EOT_EstadosOrden"]
    claves = {str(r[0]).strip() for r in ws.iter_rows(min_row=2, values_only=True)
              if r and r[0] not in (None, "")}
    h = encabezados("OT_OrdenesTrabajo")
    if "EstadoOrdenID" in h:
        i = h.index("EstadoOrdenID")
        ws2 = wb["OT_OrdenesTrabajo"]
        usados = {str(r[i]).strip() for r in ws2.iter_rows(min_row=2, values_only=True)
                  if r and len(r) > i and r[i] not in (None, "")}
        huerfanos = usados - claves
        if huerfanos:
            falla("F-08", "OT_OrdenesTrabajo.EstadoOrdenID usa %s, que no existe en "
                          "EOT_EstadosOrden" % sorted(huerfanos))
        else:
            oks.append("Los %d estados distintos usados por las ordenes resuelven contra "
                       "EOT_EstadosOrden" % len(usados))

# ------------------------- F-09 la asignacion de zona resuelve por los dos lados
def _claves(hoja, col=0):
    ws = wb[hoja]
    return {str(r[col]).strip().replace(".0", "") for r in ws.iter_rows(min_row=2, values_only=True)
            if r and r[col] not in (None, "")}


def _usados(hoja, columna):
    h = encabezados(hoja)
    if columna not in h:
        return None
    i = h.index(columna)
    ws = wb[hoja]
    return {str(r[i]).strip().replace(".0", "") for r in ws.iter_rows(min_row=2, values_only=True)
            if r and len(r) > i and r[i] not in (None, "")}


if "ASG_AsignacionZona" in wb.sheetnames:
    for columna, destino in (("UsuarioID", "USR_Usuarios"),
                             ("UnidadFuncionalID", "UNF_UnidadesFuncionales")):
        usados = _usados("ASG_AsignacionZona", columna)
        if usados is None:
            falla("F-09", "ASG_AsignacionZona.%s no existe" % columna)
            continue
        huerfanos = usados - _claves(destino)
        if huerfanos:
            falla("F-09", "ASG_AsignacionZona.%s usa %s, que no existe en %s. El Security Filter "
                          "devolveria cero activos para esos usuarios"
                  % (columna, sorted(huerfanos), destino))
        else:
            oks.append("ASG_AsignacionZona.%s resuelve contra %s" % (columna, destino))

# ---------------------- F-10 los hijos resuelven contra sus padres
# El fallo que esta comprobacion habria atrapado: CHK_Checklists.OTID se renombro
# a MantenimientoID y su fila siguio guardando 'OT-0001', que es una ORDEN.
# Renombrar un encabezado no cambia lo que el dato significa.
CADENA = [
    ("CHK_Checklists", "MantenimientoID", "MAN_Mantenimientos"),
    ("CHD_ChecklistDetalle", "ChecklistID", "CHK_Checklists"),
    ("FOT_Fotografias", "MantenimientoID", "MAN_Mantenimientos"),
    ("FIR_Firmas", "MantenimientoID", "MAN_Mantenimientos"),
    ("MAN_Mantenimientos", "OTID", "OT_OrdenesTrabajo"),
    ("OT_OrdenesTrabajo", "ActivoID", "ACT_Activos"),
    ("LST_ValoresLista", "PreguntaID", "FRM_Preguntas"),
    ("TIP_TiposActivo", "FormularioID", "FRM_Formularios"),
]
for tabla, columna, destino in CADENA:
    if tabla not in wb.sheetnames or destino not in wb.sheetnames:
        continue
    usados = _usados(tabla, columna)
    if usados is None:
        falla("F-10", "%s.%s no existe" % (tabla, columna))
        continue
    if not usados:
        continue                      # tabla vacia: nada que resolver todavia
    huerfanos = usados - _claves(destino)
    if huerfanos:
        falla("F-10", "%s.%s guarda %s, que no existe en %s. Al convertir a Ref esas filas "
                      "quedan huerfanas y AppSheet no lo anuncia"
              % (tabla, columna, sorted(huerfanos)[:5], destino))
    else:
        oks.append("%s.%s resuelve contra %s (%d valores)"
                   % (tabla, columna, destino, len(usados)))

# ------------- F-11 las listas de claves dicen la verdad sobre la hoja
# CLAVE_LEGIBLE decide cuando V-17 permite comparar un Ref contra un literal.
# Si la lista y la hoja divergen, V-17 miente en una de las dos direcciones: o
# bloquea trabajo correcto, o deja pasar el defecto que existe para cazar.
#
# La comprobacion es ASIMETRICA a proposito:
#   - FALLO si una tabla FUERA de CLAVE_LEGIBLE tiene clave de texto legible.
#     Ese es el caso que bloquea trabajo correcto.
#   - AVISO si una tabla DENTRO la tiene numerica. Molesta, no rompe.
#   - CLAVE_GENERADA queda exenta de las dos: sus valores parecen legibles hoy
#     solo por el fixture de la Fase A, y seran aleatorios en cuanto la
#     aplicacion cree la primera fila.
def _clave_es_legible(hoja):
    """Legible solo si TODAS las claves lo son.

    Mirar solo la primera fila hacia que el resultado dependiera del ORDEN de la
    hoja, que AppSheet no garantiza: reordenando USR_Usuarios para que la clave
    '3aa202ee' quedara primera, F-11 exigia meterla en CLAVE_LEGIBLE. Y esa
    exigencia es falsa -su clave es un surrogate numerico 2..11 con una anomalia-
    y habria apagado V-17 para seis referencias.
    """
    ws = wb[hoja]
    claves = [r[0] for r in ws.iter_rows(min_row=2, values_only=True)
              if r and r[0] not in (None, "")]
    if not claves:
        return None                  # tabla vacia: no se puede decidir
    return all(isinstance(v, str) and not v.replace(".", "").replace("-", "").isdigit()
               for v in claves)


for tabla in MODELO:
    if tabla not in wb.sheetnames or tabla in CLAVE_GENERADA:
        continue
    legible = _clave_es_legible(tabla)
    if legible is None:
        aviso("F-11", "%s esta vacia en la hoja: no se puede decidir si su clave es legible. "
                      "La comprobacion se salta, y conviene saberlo" % tabla)
        continue
    if legible and tabla not in CLAVE_LEGIBLE:
        falla("F-11", "%s tiene clave de texto legible en la hoja y NO esta en CLAVE_LEGIBLE. "
                      "V-17 bloqueara una comparacion correcta contra ella" % tabla)
    elif not legible and tabla in CLAVE_LEGIBLE:
        aviso("F-11", "%s esta en CLAVE_LEGIBLE pero su clave en la hoja es numerica. "
                      "Sacala, o V-17 dejara pasar un defecto real" % tabla)
    elif legible:
        oks.append("%s: clave legible, coherente con CLAVE_LEGIBLE" % tabla)

for tabla in CLAVE_GENERADA:
    if tabla in wb.sheetnames and tabla in CLAVE_LEGIBLE:
        falla("F-11", "%s esta en CLAVE_GENERADA y tambien en CLAVE_LEGIBLE. Su clave sera "
                      "aleatoria en cuanto la app cree una fila: no puede estar en las dos" % tabla)

# ------------- F-12 el dato no puede contradecir a la regla que lo calcula
# RG-19 calcula CierreConExcepcion como [Precision_GPS] > 50 (umbral D-04). Si
# el dato de la hoja dice otra cosa, la regla lo SOBREESCRIBE en cuanto alguien
# toque la fila, porque es una App formula. Y con MotivoExcepcion ya escrito, la
# fila acaba diciendo dos cosas.
#
# Es la misma forma que tuvo el defecto de RG-16: una regla cuyo dato la
# desmiente, y ninguna comprobacion que lo viera.
# El umbral NO se codifica aqui: vive en PAR_Parametros y el administrador lo
# ajusta con las pruebas de campo. Este script lee el de la hoja si existe, y si
# no cae al declarado en el modelo. Que la hoja y el modelo digan cosas distintas
# es en si un fallo.
UMBRAL_GPS = PARAMETROS["UMBRAL_GPS"][0]

if "PAR_Parametros" in wb.sheetnames:
    h = encabezados("PAR_Parametros")
    if "ParametroID" in h and "Valor" in h:
        iv = h.index("Valor")
        for r in wb["PAR_Parametros"].iter_rows(min_row=2, values_only=True):
            if not r or r[0] in (None, ""):
                continue
            clave = str(r[0]).strip()
            if clave not in PARAMETROS:
                aviso("F-13", "PAR_Parametros tiene '%s', que el modelo no declara" % clave)
                continue
            esperado = PARAMETROS[clave][0]
            actual = r[iv]
            # El umbral solo se adopta si es un numero: adoptarlo antes de
            # comprobarlo hacia que F-12 comparase float contra str y reventara.
            if clave == "UMBRAL_GPS" and isinstance(actual, (int, float))                     and not isinstance(actual, bool):
                UMBRAL_GPS = actual
            if actual is None:
                falla("F-13", "PAR_Parametros '%s' esta vacio. Las reglas que lo leen no "
                              "resolveran" % clave)
                continue
            # El TIPO importa tanto como el valor: si la celda guarda "40" como
            # texto, LOOKUP() devuelve una cadena y [Precision_GPS] > "40" no es
            # la misma comparacion. Y un texto no numerico reventaba este script
            # con un traceback en vez de reportar el fallo.
            if not isinstance(actual, (int, float)) or isinstance(actual, bool):
                falla("F-13", "PAR_Parametros '%s' guarda %r, que no es un numero sino %s. "
                              "LOOKUP() devolvera texto y la comparacion de la regla no operara"
                      % (clave, actual, type(actual).__name__))
                continue
            if float(actual) != float(esperado):
                aviso("F-13", "PAR_Parametros '%s' vale %s en la hoja y %s en el modelo. Si es un "
                              "ajuste de campo, actualiza PARAMETROS en modelo_objetivo.py"
                      % (clave, actual, esperado))
        faltan = set(PARAMETROS) - {str(r[0]).strip() for r in
                                    wb["PAR_Parametros"].iter_rows(min_row=2, values_only=True)
                                    if r and r[0] not in (None, "")}
        if faltan:
            falla("F-13", "PAR_Parametros no tiene %s. RG-19 y RG-01 los leen con LOOKUP()"
                  % sorted(faltan))
        else:
            oks.append("PAR_Parametros tiene los %d parametros que declara el modelo" % len(PARAMETROS))

if "MAN_Mantenimientos" in wb.sheetnames:
    h = encabezados("MAN_Mantenimientos")
    if "Precision_GPS" in h and "CierreConExcepcion" in h:
        ip, ic = h.index("Precision_GPS"), h.index("CierreConExcepcion")
        ws = wb["MAN_Mantenimientos"]
        revisadas = 0
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or r[0] in (None, "") or len(r) <= max(ip, ic):
                continue
            precision, marca = r[ip], r[ic]
            if precision is None or marca is None:
                continue
            revisadas += 1
            esperado = precision > UMBRAL_GPS
            actual = str(marca).strip().upper() in ("TRUE", "VERDADERO", "1", "SI")
            if esperado != actual:
                falla("F-12", "MAN_Mantenimientos '%s': Precision_GPS=%s, luego RG-19 calculara "
                              "CierreConExcepcion=%s, pero la hoja dice %s. La App formula "
                              "sobreescribira el dato en cuanto alguien toque la fila"
                      % (r[0], precision, esperado, actual))
        if revisadas:
            oks.append("Las %d filas de MAN_Mantenimientos coinciden con RG-19 (umbral %d m)"
                       % (revisadas, UMBRAL_GPS))

# ------------- F-14 OT_OrdenesTrabajo y EST_Activo tienen Activo = TRUE
for tabla in ("OT_OrdenesTrabajo", "EST_Activo"):
    if tabla in wb.sheetnames:
        h = encabezados(tabla)
        if "Activo" in h:
            idx = h.index("Activo")
            ws = wb[tabla]
            falsos = 0
            total = 0
            for r in ws.iter_rows(min_row=2, values_only=True):
                if not r or r[0] in (None, ""): continue
                total += 1
                val = r[idx]
                if str(val).strip().upper() not in ("TRUE", "VERDADERO", "1", "SI"):
                    falsos += 1
            if falsos > 0:
                falla("F-14", "%s.Activo tiene %d filas sin TRUE. Deben estar activas." % (tabla, falsos))
            else:
                oks.append("%s.Activo tiene TRUE en sus %d filas" % (tabla, total))

# ------------- F-15 ACT_Activos fila 34 esta dado de baja
if "ACT_Activos" in wb.sheetnames:
    h = encabezados("ACT_Activos")
    if "ActivoID" in h and "Activo" in h:
        id_idx = h.index("ActivoID")
        act_idx = h.index("Activo")
        ws = wb["ACT_Activos"]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or r[0] in (None, ""): continue
            if str(r[id_idx]).strip() == "34":
                val = r[act_idx]
                if str(val).strip().upper() in ("TRUE", "VERDADERO", "1", "SI"):
                    falla("F-15", "ACT_Activos fila 34 dice Activo=TRUE pero esta Retirado. Debe ser FALSE.")
                else:
                    oks.append("ACT_Activos fila 34 (SUBE-001) tiene Activo=FALSE correctamente")

# ------------- F-19 lo declarado como retirado sigue existiendo en la hoja
#
# CAMPOS_RETIRADOS dice "esta columna existe en la hoja y el modelo no la usa".
# Si la columna YA NO ESTA, la entrada esta obsoleta y contamina todo lo que se
# genere de ella: el 2026-08-09 dos entradas fantasma -ACT_Activos.SedeID y
# MAN_Mantenimientos.ActivoID- llegaron a un prompt de ejecucion, y el agente que
# lo seguia perdio tiempo buscandolas en el editor.
#
# Es la misma familia que F-18: el modelo dice una cosa, el archivo otra, y
# nadie los cruzaba.
#
# AMPLIADA el 2026-08-10, al aparecer una segunda clase de libro. Hay dos
# estados sanos y uno peligroso, y antes solo se contemplaba el primero:
#
#   estan las 43   la hoja heredada. Se retiran ocultandolas, a proposito
#   no esta ninguna la hoja limpia generada del modelo. Es el objetivo
#   estan algunas  alguien borro parte a mano y dejo el resto. ESTE es el malo
#
# El estado mixto es el que hace dano: la documentacion generada manda ocultar
# columnas que ya no existen, y quien lo ejecuta las busca en vano, mientras las
# que si quedan siguen apareciendo en el formulario del tecnico.
fantasma = []
presente = []
for tabla, campos in CAMPOS_RETIRADOS.items():
    if tabla not in wb.sheetnames:
        continue
    presentes = set(encabezados(tabla))
    for campo in campos:
        (presente if campo in presentes else fantasma).append("%s.%s" % (tabla, campo))

if fantasma and not presente:
    oks.append("Hoja limpia: ninguna de las %d columnas retiradas existe ya. "
               "No hay nada que ocultar" % len(fantasma))
elif fantasma:
    avisos.append("[F-19] ESTADO MIXTO: de las %d columnas retiradas quedan %d en la hoja y "
                  "%d ya no estan (%s). O se conservan todas -y se ocultan- o no queda "
                  "ninguna. A medias, la documentacion generada manda ocultar columnas "
                  "inexistentes" % (len(fantasma) + len(presente), len(presente),
                                    len(fantasma), ", ".join(sorted(fantasma))))
else:
    oks.append("Toda columna declarada como retirada sigue presente en la hoja")

# ------------- F-20 la clave no mezcla numeros y texto
#
# AppSheet TIPA LA COLUMNA CLAVE SEGUN LA MAYORIA de sus valores. Si diez son
# numeros y uno es texto, la tipa Number y DESCARTA esa fila: sin error, sin
# aviso y sin que aparezca en ninguna parte. Paso el 2026-08-10 con
# USR_Usuarios, donde una clave alfanumerica -3aa202ee, generada por el propio
# AppSheet con UNIQUEID- dejaba a un tecnico fuera del sistema. Se descubrio
# porque la API devolvia 10 usuarios y la hoja tenia 11; ningun verificador lo
# miraba.
#
# El modelo declara TODAS las claves como Text, asi que una clave enteramente
# numerica tampoco es inocente: AppSheet la tipara Number, y el dia que alguien
# cree una fila desde la aplicacion con UNIQUEID, esa fila se pierde igual.
import re as _re


def _es_numero(v):
    return bool(_re.fullmatch(r"-?\d+(\.\d+)?", str(v).strip()))


for tabla in MODELO:
    if tabla not in wb.sheetnames:
        continue
    ws = wb[tabla]
    valores = [r[0] for r in ws.iter_rows(min_row=2, values_only=True)
               if r and r[0] not in (None, "")]
    if not valores:
        continue
    numeros = sum(1 for v in valores if _es_numero(v))
    textos = len(valores) - numeros
    if numeros and textos:
        minoria = ([str(v) for v in valores if not _es_numero(v)] if numeros > textos
                   else [str(v) for v in valores if _es_numero(v)])
        falla("F-20", "%s: la clave mezcla %d numericas y %d de texto. AppSheet la tipa por la "
                      "mayoria y DESCARTA la minoria sin decir nada. Se pierden: %s"
              % (tabla, numeros, textos, ", ".join(sorted(minoria)[:8])))
    elif numeros:
        aviso("F-20", "%s: la clave es enteramente numerica y el modelo la declara Text. "
                      "AppSheet la tipara Number, y una fila creada desde la aplicacion con "
                      "UNIQUEID se perderia. Forzar Text en el editor" % tabla)
    else:
        oks.append("%s: la clave es de texto en sus %d filas" % (tabla, len(valores)))

# ------------- F-18 ninguna pestana del modelo esta OCULTA
#
# AppSheet IGNORA las pestanas ocultas al escanear un libro. No avisa: la tabla
# simplemente no aparece en Add data, y quien busque ROL_Roles en el desplegable
# de Source table no la encontrara nunca.
#
# openpyxl SI las lee, y por eso este script cerro la Fase A dos veces sobre un
# libro con ACT_Activos, USR_Usuarios y TIP_TiposActivo ocultas. Vio las 32 y
# dijo CERRADA mientras AppSheet solo veia 24.
#
# Es el mismo modo de fallo de F-17: la comprobacion pasaba porque medía lo que
# no era. Se detecta con sheet_state, que openpyxl expone y nadie miraba.
ocultas = [h for h in wb.sheetnames if wb[h].sheet_state != "visible"]
del_modelo = [h for h in ocultas if h in MODELO]
if del_modelo:
    falla("F-18", "%d pestanas del modelo estan OCULTAS y AppSheet no las vera: %s. "
                  "Mostrarlas en Google Sheets antes de dar de alta las tablas"
                  % (len(del_modelo), ", ".join(sorted(del_modelo))))
elif ocultas:
    avisos.append("[F-18] %d pestanas ocultas, ninguna del modelo: %s"
                  % (len(ocultas), ", ".join(sorted(ocultas))))
else:
    oks.append("Ninguna de las %d pestanas esta oculta" % len(wb.sheetnames))

# ------- F-16 la clave y quien la apunta se guardan en el mismo formato
# Un Ref guarda el VALOR de la clave. Si la clave esta como numero (2.0) y quien
# la apunta como texto ('2'), al tipar ambas AppSheet tiene que convertir, y de
# como convierta depende que la referencia resuelva. No es hipotetico: es la
# clase de huerfano silencioso que este proyecto lleva meses arrastrando.
def _tipos_de(hoja, columna):
    h = encabezados(hoja)
    if columna not in h:
        return None
    i = h.index(columna)
    ws = wb[hoja]
    t = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r and len(r) > i and r[i] not in (None, ""):
            t.add("numero" if isinstance(r[i], (int, float)) and not isinstance(r[i], bool)
                  else type(r[i]).__name__)
    return t


for tabla, d in MODELO.items():
    if tabla not in wb.sheetnames:
        continue
    for c in d["columnas"]:
        if c["tipo"] != "Ref" or not c.get("ref") or c["ref"] not in wb.sheetnames:
            continue
        destino = c["ref"]
        pk = next((x["nombre"] for x in MODELO[destino]["columnas"] if x.get("pk")), None)
        if not pk:
            continue
        t_ref, t_pk = _tipos_de(tabla, c["nombre"]), _tipos_de(destino, pk)
        if not t_ref or not t_pk:
            continue
        if len(t_ref) > 1:
            falla("F-16", "%s.%s mezcla %s en la misma columna. Al tipar como Ref, unas filas "
                          "resolveran y otras no" % (tabla, c["nombre"], sorted(t_ref)))
        elif t_ref != t_pk:
            # Si lo mezclado es la CLAVE del destino y todos los valores que la
            # apuntan resuelven, no hay ninguna fila rota: es una anomalia de la
            # clave, no de esta referencia. Fallar aqui era contar cinco veces el
            # mismo defecto -una celda de USR_Usuarios- y bloquear ademas
            # referencias aplazadas a ESPEC-003.
            usados = _usados(tabla, c["nombre"]) or set()
            huerfanos = usados - _claves(destino)
            if len(t_pk) > 1 and not huerfanos:
                aviso("F-16", "%s.%s guarda %s y la clave %s.%s esta mezclada (%s), pero las %d "
                              "filas resuelven. Se arregla en la clave, no aqui"
                      % (tabla, c["nombre"], sorted(t_ref), destino, pk, sorted(t_pk), len(usados)))
            else:
                falla("F-16", "%s.%s guarda %s y la clave %s.%s guarda %s. Un Ref guarda el valor "
                              "de la clave: si no comparten formato, la conversion decide si "
                              "resuelve" % (tabla, c["nombre"], sorted(t_ref), destino, pk,
                                            sorted(t_pk)))

# ------- F-17 ninguna columna que AppSheet vaya a gestionar contiene formulas
# Una formula de la hoja produce un valor que AppSheet puede leer, pero que
# cambia solo si cambia su origen. TIP_TiposActivo.FormularioID sale de
# =CONCAT("FRM_",MID(B2,1,4)): renombrar un tipo de activo repunta en silencio
# el checklist que abre la aplicacion.
for hoja in wb_formulas.sheetnames:
    if hoja not in MODELO:
        continue
    h = encabezados(hoja)
    ws_f = wb_formulas[hoja]
    for i, nombre in enumerate(h):
        formulas = 0
        for r in ws_f.iter_rows(min_row=2, values_only=True):
            if r and len(r) > i and isinstance(r[i], str) and r[i].startswith("="):
                formulas += 1
        if not formulas:
            continue
        # El valor cacheado es lo que hay que escribir literalmente al sustituir
        # la formula. Sin el, quien la borre se queda sin el dato.
        primero = next((r[i] for r in wb[hoja].iter_rows(min_row=2, values_only=True)
                        if r and len(r) > i and r[i] not in (None, "")), None)
        falla("F-17", "%s.%s contiene %d formulas de hoja de calculo. Su valor cambia si cambia el "
                      "origen, y AppSheet lo sobreescribe al escribir la fila. Primer valor "
                      "calculado: %r" % (hoja, nombre, formulas, primero))

# ------------------------------------------------------------------- informe
print("=" * 78)
print("VERIFICACION DE LA FASE A")
print("=" * 78)
print("Archivo: %s" % os.path.basename(ruta))
print("Hojas:   %d" % len(wb.sheetnames))
print("-" * 78)
if oks:
    print("CONFORMES (%d):" % len(oks))
    for o in oks:
        print("  ok", o)
    print()
if fallos:
    print("FALLOS (%d) — la Fase A NO esta cerrada:" % len(fallos))
    for f in fallos:
        print("  x", f)
    print()
if avisos:
    print("AVISOS (%d) — esperados, no bloquean:" % len(avisos))
    for a in avisos:
        print("  -", a)
    print()
print("=" * 78)
print("FASE A CERRADA" if not fallos else "FASE A INCOMPLETA: faltan %d puntos" % len(fallos))
sys.exit(1 if fallos else 0)
