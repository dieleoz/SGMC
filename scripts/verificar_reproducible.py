# -*- coding: utf-8 -*-
"""Comprueba que generar la plantilla dos veces da el mismo archivo.

El quinto verificador de repositorio, y nace de un fallo que los otros cuatro
no podian ver.

Que paso
--------
El 2026-08-10, al resembrar las claves como alfanumericas, el generador dejo de
ser idempotente sin que nada protestara. La garantia de catalogo buscaba las
filas POR CLAVE; la resiembra cambiaba esa clave; en la pasada siguiente ya no
las encontraba y las volvia a anadir. SED_Sedes acabo con las seis
edificaciones DUPLICADAS y CAL_Calzadas con dos Separador, y cada ejecucion
habria anadido seis mas.

Y paso los cuatro verificadores: el modelo era coherente, la Fase A cerraba, la
prosa cuadraba y los enlaces resolvian. Ninguno mira si el generador se
contradice consigo mismo, porque todos miran UN archivo. El defecto solo existe
ENTRE DOS EJECUCIONES.

Por que importa
---------------
Un generador no idempotente convierte su salida en un artefacto que hay que
conservar, porque rehacerlo da algo distinto. Es exactamente lo que este
proyecto decidio no volver a tener.

Uso:  python scripts/verificar_reproducible.py
Sale con codigo 1 si dos ejecuciones seguidas no coinciden.
"""
import os
import shutil
import subprocess
import sys
import tempfile

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "scripts"))

try:
    import openpyxl
except ImportError:
    print("Falta openpyxl."); sys.exit(2)

from sistema import VOLCADO

DESTINO = os.path.join(RAIZ, VOLCADO)
GENERADOR = os.path.join(RAIZ, "scripts", "generar_plantilla.py")


def contenido(ruta):
    """Todas las celdas de todas las pestanas. Sin metadatos: el .xlsx guarda la
    fecha de modificacion, asi que comparar los bytes daria distinto siempre."""
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    try:
        return {h: [r for r in wb[h].iter_rows(values_only=True)]
                for h in wb.sheetnames}
    finally:
        # En modo read_only openpyxl deja el archivo ABIERTO hasta que se cierra
        # a mano. En Windows eso hace que TemporaryDirectory reviente al limpiar
        # -WinError 32, el proceso no tiene acceso al archivo- y el script sale
        # con codigo 1 SIN HABER COMPARADO NADA.
        #
        # Es exactamente la trampa que este proyecto ya piso el 2026-08-11 con
        # verificar_datos.py: un verificador que sale con 1 por un crash parece
        # una deteccion, y se lee como si hubiera cazado algo. Peor aun al reves:
        # aqui se leia como que la reproducibilidad fallaba, cuando ni siquiera
        # se habia llegado a comprobar.
        wb.close()


ancho = "=" * 78
print(ancho)
print("REPRODUCIBILIDAD DEL GENERADOR")
print(ancho)
print("Se genera dos veces seguidas y se compara celda a celda.")
print("")

if not os.path.exists(DESTINO):
    print("No existe %s. Genera la plantilla primero." % VOLCADO)
    sys.exit(2)

with tempfile.TemporaryDirectory() as tmp:
    # La primera pasada parte del archivo actual, que es lo que hace el uso real.
    subprocess.run([sys.executable, GENERADOR], capture_output=True, cwd=RAIZ)
    primera = os.path.join(tmp, "primera.xlsx")
    shutil.copy(DESTINO, primera)

    subprocess.run([sys.executable, GENERADOR], capture_output=True, cwd=RAIZ)
    a, b = contenido(primera), contenido(DESTINO)

hojas_a, hojas_b = set(a), set(b)
fallos = []

if hojas_a != hojas_b:
    fallos.append("Las pestanas no coinciden: sobran %s, faltan %s"
                  % (sorted(hojas_b - hojas_a) or "ninguna",
                     sorted(hojas_a - hojas_b) or "ninguna"))

print("%-26s %10s %10s  %s" % ("PESTANA", "1a VEZ", "2a VEZ", ""))
for h in sorted(hojas_a & hojas_b):
    na, nb = len(a[h]), len(b[h])
    if a[h] == b[h]:
        continue
    nota = "CRECE %+d filas" % (nb - na) if na != nb else "mismas filas, distinto contenido"
    fallos.append("%s: la segunda ejecucion da algo distinto. %s" % (h, nota))
    print("%-26s %10d %10d  %s" % (h, na, nb, nota))

print("")
print(ancho)
if fallos:
    for f in fallos:
        print("  x %s" % f)
    print(ancho)
    print("EL GENERADOR NO ES REPRODUCIBLE")
    print("Su salida deja de ser algo que se rehace y pasa a ser algo que hay que conservar.")
    print(ancho)
    sys.exit(1)

print("REPRODUCIBLE: las %d pestanas salen identicas" % len(hojas_a))
print(ancho)
