# -*- coding: utf-8 -*-
"""Comprueba que docs/SISTEMA.md sigue siendo verdad.

Por que existe
--------------
`docs/SISTEMA.md` describe el sistema en presente: 28 tablas, 39 referencias, el
radio de geofencing por tipo, cuantas preguntas son borrador. Nada de eso es una
opinion: **todo sale de un comando**.

Y por eso no necesita pasar por el pipeline. Una ESPEC propone un cambio y hace
falta que alguien la refute; un documento que solo describe no se refuta, **se
comprueba**. Este script es esa comprobacion.

El problema que resuelve
------------------------
Este proyecto lleva seis cifras publicadas que envejecieron en silencio -113,
205, 52, 39/39, 36 y «21 reglas»- porque nadie volvia a mirarlas. Un documento
de referencia sin verificador es una promesa de que alguien se acordara, y no se
acuerda nadie.

Con esto, SISTEMA.md **se desmiente solo** el dia que deje de ser cierto.

Que NO comprueba
----------------
Lo que ningun comando puede ver: los tipos de columna del editor, las
expresiones, los permisos y el Label. Eso lo dice el propio documento en su
seccion 5, y `scripts/lectura_de_vuelta.py` lo declara por clase de cambio.

Uso:  python scripts/verificar_sistema.py
Sale con codigo 1 si el documento afirma algo que ya no es cierto.
"""
import os
import re
import sys

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "scripts"))

DOC = os.path.join(RAIZ, "docs", "SISTEMA.md")

if not os.path.exists(DOC):
    print("No existe docs/SISTEMA.md."); sys.exit(2)

texto = open(DOC, encoding="utf-8").read()

from modelo_objetivo import MODELO, REGLAS, RETIRADAS, CLAVE_GENERADA
from sistema import APP_NOMBRE, HOJA_NOMBRE

fallos, comprobadas = [], []


def afirma(que, esperado, como, patron=None):
    """Comprueba que el documento dice `esperado` SOBRE `que`.

    El `patron` importa. La primera version solo miraba si el numero aparecia
    en algun sitio del documento, y con eso NO CAZABA NADA: cambiar «23 reglas»
    por «21 reglas» pasaba en verde, porque «23» seguia estando en otra frase.

    Un verificador que pasa por el motivo equivocado es peor que no tenerlo:
    ocupa el sitio del que si comprobaria. Se descubrio probandolo, que es la
    unica forma.
    """
    comprobadas.append(que)
    pat = patron or (r"%s\s+%s" % (esperado, que.split()[0]))
    if not re.search(pat, texto, re.I):
        fallos.append("%s: el documento no dice «%s» donde deberia. Lo derivado "
                      "hoy es %s (%s)" % (que, esperado, esperado, como))


def no_afirma(que, prohibido, motivo):
    """Comprueba que el documento NO dice algo que ya es falso."""
    comprobadas.append(que)
    if prohibido in texto:
        fallos.append("%s: el documento todavia dice «%s», y %s"
                      % (que, prohibido, motivo))


ancho = "=" * 78
print(ancho)
print("docs/SISTEMA.md SIGUE SIENDO VERDAD?")
print(ancho)
print("")

# ---------------------------------------------------- las cifras del modelo
tablas = len(MODELO)
columnas = sum(len(MODELO[t]["columnas"]) for t in MODELO)
refs = sum(1 for t in MODELO for c in MODELO[t]["columnas"] if c.get("ref"))

afirma("tablas", tablas, "len(MODELO)", r"%d\s+tablas" % tablas)
afirma("columnas", columnas, "suma de columnas de MODELO", r"%d\s+columnas" % columnas)
afirma("referencias", refs, "columnas con ref", r"%d\s+referencias" % refs)
afirma("reglas", len(REGLAS), "len(REGLAS)", r"%d\s+reglas" % len(REGLAS))
afirma("tablas retiradas", len(RETIRADAS), "len(RETIRADAS)",
       r"(cinco|%d)\s+tablas se declaran retiradas" % len(RETIRADAS))

# ------------------------------------------------- los identificadores vivos
afirma("nombre de la aplicacion", APP_NOMBRE, "scripts/sistema.py", re.escape(APP_NOMBRE))
afirma("nombre de la hoja", HOJA_NOMBRE, "scripts/sistema.py", re.escape(HOJA_NOMBRE))

# ------------------------------------------- el radio de geofencing por tipo
try:
    from catalogo_tipos import TIPOS_ACTIVO
    radios = sorted({t[7] for t in TIPOS_ACTIVO if t[7]})
    afirma("tipos de activo", len(TIPOS_ACTIVO), "len(TIPOS_ACTIVO)",
           r"%d\s+(tipos|filas)" % len(TIPOS_ACTIVO))
    for r in radios:
        # el documento los escribe con coma decimal
        if str(r).replace(".", ",") not in texto:
            fallos.append("radio %s: el documento no lo menciona, y esta en "
                          "catalogo_tipos.py" % r)
    comprobadas.append("radios de geofencing")
except ImportError:
    pass

# ---------------------------------------- lo que se dice de la hoja de datos
try:
    import openpyxl
    from sistema import VOLCADO
    wb = openpyxl.load_workbook(os.path.join(RAIZ, VOLCADO), read_only=True)
    if "FRM_Preguntas" in wb.sheetnames:
        ws = wb["FRM_Preguntas"]
        cab = [str(c.value) for c in next(ws.iter_rows(max_row=1))]
        if "Ayuda" in cab:
            i = cab.index("Ayuda") + 1
            total = ws.max_row - 1
            borrador = sum(1 for r in range(2, ws.max_row + 1)
                           if "BORRADOR" in str(ws.cell(r, i).value or ""))
            afirma("preguntas del banco", total, "filas de FRM_Preguntas", r"%d\s+preguntas" % total)
            afirma("preguntas en borrador", borrador, "las que llevan la marca",
                   r"%d\s+de\s+las\s+%d" % (borrador, total))
except Exception as e:
    print("  ! no se pudo leer la hoja: %s" % e)
    print("")

# --------------------------------- afirmaciones que dejarian de ser ciertas
#
# No basta con que las cifras cuadren: el documento tambien afirma COSAS, y
# algunas dejarian de valer si el modelo cambia. Estas son las que se pueden
# comprobar mecanicamente.
if "OT_OrdenesTrabajo" in CLAVE_GENERADA:
    no_afirma("clave de las ordenes", "`OTID` es legible",
              "OT_OrdenesTrabajo esta en CLAVE_GENERADA: su clave la genera UNIQUEID()")

virtuales = [r for r in REGLAS
             if r["tipo"] == "App formula" and r.get("columna") == "(tabla)"]
# El documento las NOMBRA en vez de contarlas, y esta mejor asi: una lista de
# dos tablas envejece de forma visible; un «dos» no. La comprobacion se ajusta
# al documento, no al reves.
for _r in virtuales:
    comprobadas.append("columna virtual de %s" % _r["tabla"])
    if _r["tabla"] not in texto:
        fallos.append("columna virtual de %s: la regla %s la declara y el "
                      "documento no la nombra" % (_r["tabla"], _r["id"]))

# ---------------------------------------------------------------- salida
for f in fallos:
    print("  x %s" % f)

print("")
print(ancho)
if fallos:
    print("SISTEMA.md AFIRMA %d COSAS QUE YA NO SON CIERTAS" % len(fallos))
    print("")
    print("No lo arregles a mano: es el documento de referencia, y una cifra")
    print("corregida a mano vuelve a envejecer. Mira que cambio en el modelo y")
    print("reescribe la frase entera con lo que hoy es verdad.")
    print(ancho)
    sys.exit(1)

print("SIGUE SIENDO VERDAD: %d afirmaciones comprobadas" % len(comprobadas))
print("")
print("Lo que este script NO puede comprobar es lo que el propio documento")
print("declara en su seccion 5: los tipos de columna, las expresiones, los")
print("permisos y el Label. Eso no lo devuelve la API, y se mira en el editor.")
print(ancho)
