import sys
import openpyxl

excel_file = r"d:\@Proyect\Sisga\BD\Modelo de Datos (4).xlsx"

RENOMBRADOS = {
  'OT_OrdenesTrabajo': [
    ['Activo',            'ActivoID'],
    ['Numero_OT',         'OTID'],
    ['Tecnico',           'TecnicoID'],
    ['SupervidorID',      'SupervisorID'],
    ['Fecha Programada',  'FechaProgramada'],
    ['Estado',            'EstadoOrdenID'],
    ['Fecha_Cierre',      'FechaCierre'],
    ['Cerrada_Por',       'CerradaPor']
  ],
  'MAN_Mantenimientos': [
    ['MttoID',                  'MantenimientoID'],
    ['Tecnico_Asignado',        'TecnicoID'],
    ['Fecha_Hora_Inicio',       'FechaHoraInicio'],
    ['Fecha_Hora_Fin',          'FechaHoraFin'],
    ['Requiere_Segunda_Visita', 'RequiereSegundaVisita'],
    ['Motivo_Pendiente',        'MotivoPendienteID'],
    ['Aprobado_Supervisor',     'AprobadoSupervisor'],
    ['Usuario_Registro',        'UsuarioRegistro'],
    ['Fecha_Hora_Registro',     'FechaHoraRegistro']
  ],
  'ACT_Activos': [
    ['EstadoID', 'EstadoActivoID'],
    ['SedeID',   'UnidadFuncionalID']
  ],
  'USR_Usuarios': [
    ['usuarioID', 'UsuarioID'],
    ['Estado',    'Activo']
  ],
  'EST_Activo': [
    ['EstadoID', 'EstadoActivoID']
  ],
  'CHK_Checklists': [
    ['OTID', 'MantenimientoID']
  ],
  'CHD_ChecklistDetalle': [
    ['Observaciones', 'Observacion']
  ]
}

COLUMNAS_NUEVAS = {
  'OT_OrdenesTrabajo': ['OTOrigenID', 'Activo'],
  'MAN_Mantenimientos': [
    'OrigenApertura', 'UbicacionEscaneo', 'FechaHoraEscaneo', 'EstadoActivoID',
    'CierreConExcepcion', 'MotivoExcepcion', 'ModoFallaID', 'FechaAprobacion',
    'ObservacionRechazo'
  ],
  'TIP_TiposActivo': ['RadioGeofencingKm'],
  'CHK_Checklists':  ['VersionFormulario'],
  'EST_Activo':      ['GeneraAlerta', 'Activo']
}

TABLAS_NUEVAS = [
  'UNF_UnidadesFuncionales', 'EOT_EstadosOrden', 'MOT_MotivosPendiente',
  'ASG_AsignacionZona', 'FAL_ModosFalla', 'NOV_Novedades', 'PLA_PlanMantenimiento'
]

LIMPIEZA_CHK = ['CHK001', '0356e6d7']

def main():
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
    except Exception as e:
        print(f"Error cargando excel: {e}")
        sys.exit(1)

    fallos = 0

    def indiceDe(hoja, nombre):
        if hoja.max_column == 0: return -1
        # header is row 1
        for cell in hoja[1]:
            if str(cell.value).strip() == nombre:
                return cell.column
        return -1

    for nombreHoja, renombrados in RENOMBRADOS.items():
        if nombreHoja not in wb.sheetnames:
            print(f"XX {nombreHoja} no existe.")
            fallos += 1
            continue
        hoja = wb[nombreHoja]
        for viejo, nuevo in renombrados:
            if indiceDe(hoja, nuevo) == -1:
                print(f"XX {nombreHoja}.{nuevo} NO existe.")
                fallos += 1
            elif indiceDe(hoja, viejo) != -1:
                print(f"XX {nombreHoja}.{viejo} sigue existiendo con el nombre viejo.")
                fallos += 1

    for nombreHoja, columnas in COLUMNAS_NUEVAS.items():
        if nombreHoja not in wb.sheetnames:
            continue
        hoja = wb[nombreHoja]
        for col in columnas:
            if indiceDe(hoja, col) == -1:
                print(f"XX {nombreHoja}.{col} NO existe.")
                fallos += 1

    for nombreHoja in TABLAS_NUEVAS:
        if nombreHoja not in wb.sheetnames:
            print(f"XX {nombreHoja} NO existe.")
            fallos += 1

    if 'CHK_Checklists' in wb.sheetnames:
        chk = wb['CHK_Checklists']
        ids = []
        for row in chk.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] is not None:
                ids.append(str(row[0]).strip())
        
        for id_borrar in LIMPIEZA_CHK:
            if id_borrar in ids:
                print(f"XX CHK_Checklists: '{id_borrar}' sigue ahi.")
                fallos += 1
        
        if 'd02d8a3d' not in ids:
            print(f"XX CHK_Checklists: se borro d02d8a3d, que debia CONSERVARSE.")
            fallos += 1

    if fallos == 0:
        print("VERIFICACION LIMPIA. Fase A cerrada.")
    else:
        print(f"VERIFICACION CON {fallos} FALLOS.")
        sys.exit(1)

if __name__ == '__main__':
    main()
