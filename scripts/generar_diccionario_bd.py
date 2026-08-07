# -*- coding: utf-8 -*-
"""Genera docs/bd.md: el diccionario de datos As-Built.

Lo lee del ARCHIVO, no de la memoria de nadie. Cruza lo que la hoja tiene hoy
contra lo que scripts/modelo_objetivo.py dice que debe tener, y marca cada
columna con su estado.

Uso:  python scripts/generar_diccionario_bd.py "BD/Modelo de Datos (9).xlsx"

Por que se genera y no se escribe
---------------------------------
La version anterior de bd.md decia "24 Hojas Produccion" cuando habia 32,
apuntaba al Excel (2) como maestro cuando vamos por el (9), y describia columnas
-MttoID, Tecnico_Asignado- renombradas dos dias antes. No por descuido: un
diccionario escrito a mano envejece en cuanto alguien toca la hoja, y este
proyecto ya perdio meses porque bd.md y el Excel describian modelos distintos
sin que nadie lo detectara.

Generandolo, la unica forma de que mienta es que mienta el archivo.
"""
import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from modelo_objetivo import (MODELO, RENOMBRADOS, RETIPADOS, CAMPOS_RETIRADOS,
                             RETIRADAS, PARAMETROS)

try:
    import openpyxl
except ImportError:
    print("Falta openpyxl."); sys.exit(2)

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ruta = sys.argv[1] if len(sys.argv) > 1 else "BD/Modelo de Datos (9).xlsx"
if not os.path.isabs(ruta):
    ruta = os.path.join(RAIZ, ruta)
SALIDA = os.path.join(RAIZ, "docs", "bd.md")

wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)

# ------------------------------------------------------------------ indices
# De donde viene cada columna: renombrada, retipada, retirada o del modelo.
origen = {}          # (tabla, columna) -> nota
for tabla, mapa in RENOMBRADOS.items():
    for viejo, (nuevo, motivo) in mapa.items():
        origen[(tabla, nuevo)] = "Antes `%s`. %s" % (viejo, motivo)
retipar = {(t, c): d for t, m in RETIPADOS.items() for c, d in m.items()}
retirados = {(t, c): motivo for t, campos in CAMPOS_RETIRADOS.items()
             for c, motivo in campos.items()}

L = []
w = L.append


def encabezados(hoja):
    ws = wb[hoja]
    fila = next(ws.iter_rows(min_row=1, max_row=1))
    return [str(c.value).strip() for c in fila if c.value is not None]


def filas(hoja):
    ws = wb[hoja]
    return sum(1 for r in ws.iter_rows(min_row=2, values_only=True)
               if any(v not in (None, "") for v in r))


def muestra(hoja, n=2):
    ws = wb[hoja]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r and r[0] not in (None, ""):
            out.append(str(r[0]))
        if len(out) >= n:
            break
    return ", ".join("`%s`" % x for x in out) if out else "vacía"


hojas = wb.sheetnames
n_filas = {h: filas(h) for h in hojas}

# ------------------------------------------------------------------ cabecera
w("# Diccionario de datos — As-Built")
w("")
w("**Generado automáticamente** por `scripts/generar_diccionario_bd.py` desde")
w("`%s`. No editar a mano." % os.path.basename(ruta))
w("")
w("Describe **lo que la hoja tiene hoy**, no lo que debería tener. El modelo objetivo está en")
w("`ARQUITECTURA_OBJETIVO_SGMC.md`; aquí se marca la distancia entre uno y otro.")
w("")
w("La versión anterior de este documento decía «24 hojas» cuando había 32, apuntaba a un Excel")
w("maestro de hace tres días y describía columnas renombradas. Por eso ahora se genera: **la única")
w("forma de que mienta es que mienta el archivo**.")
w("")
w("| | |")
w("|---|---|")
w("| Backend de producción | Google Sheets `1a4MmZ0u9sNgWmyiR2OPJo9YuUEKJFftbJWMW-KbITRc` |")
w("| Aplicación | AppSheet `SGMC-886843353` |")
w("| Hojas | **%d** |" % len(hojas))
w("| Filas con datos | **%d** |" % sum(n_filas.values()))
w("| Generado el | %s |" % datetime.now().strftime("%Y-%m-%d"))
w("")
w("---")
w("")

# -------------------------------------------------------------- estado global
w("## 1. Qué falta para llegar al modelo objetivo")
w("")
pendientes = [(t, c) for t, m in RETIPADOS.items() for c in m]
w("| Concepto | Cuántos | Dónde se resuelve |")
w("|---|---|---|")
w("| Columnas que siguen siendo `Text` y deben ser `Ref` | **%d** | Fase B, `ESPEC-002` |" % len(pendientes))
w("| Columnas marcadas como retiradas que siguen en la hoja | **%d** | Pasada posterior, con datos ya migrados |"
  % sum(len(c) for c in CAMPOS_RETIRADOS.values()))
w("| Tablas marcadas como retiradas que siguen en la hoja | %d | Idem |"
  % len([t for t in RETIRADAS if t in hojas]))
w("| Tablas del modelo objetivo que ya existen | %d de %d | — |"
  % (len([t for t in MODELO if t in hojas]), len(MODELO)))
w("")
w("**Ninguna de esas columnas se borra todavía a propósito.** La Fase A no borra nada: borrar es lo")
w("único que un respaldo no vuelve gratis.")
w("")

# ------------------------------------------------------------------ inventario
w("## 2. Inventario de hojas")
w("")
w("| Hoja | Columnas | Filas | En el modelo objetivo |")
w("|---|---|---|---|")
for h in hojas:
    if h in MODELO:
        estado = "Sí" + (" · **nueva**" if MODELO[h].get("nueva") else "")
    elif h in RETIRADAS:
        estado = "**Se retira.** %s" % RETIRADAS[h]
    else:
        estado = "No figura"
    w("| `%s` | %d | %d | %s |" % (h, len(encabezados(h)), n_filas[h], estado))
w("")

# ------------------------------------------------------- parametros calibrables
if "PAR_Parametros" in hojas:
    w("## 3. Parámetros calibrables")
    w("")
    w("Umbrales que el administrador ajusta **en la hoja**, sin abrir el editor de AppSheet. Un")
    w("número escondido dentro de una expresión no se puede calibrar.")
    w("")
    ws = wb["PAR_Parametros"]
    hdr = encabezados("PAR_Parametros")
    iv = hdr.index("Valor") if "Valor" in hdr else 2
    iu = hdr.index("Unidad") if "Unidad" in hdr else 3
    w("| Parámetro | Valor en la hoja | Unidad | Declarado en el modelo | Quién lo lee |")
    w("|---|---|---|---|---|")
    lectores = {"UMBRAL_GPS": "RG-19", "RADIO_GEOFENCING_KM": "RG-01",
                "DISTANCIA_ESCANEO_CIERRE_KM": "RG-13"}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or r[0] in (None, ""):
            continue
        clave = str(r[0]).strip()
        decl = PARAMETROS.get(clave, ("—",))[0]
        w("| `%s` | %s | %s | %s | %s |"
          % (clave, r[iv], r[iu] or "", decl, lectores.get(clave, "—")))
    w("")

# ------------------------------------------------------------------ por tabla
w("## 4. Detalle por hoja")
w("")
w("Leyenda del estado de cada columna:")
w("")
w("- **Pendiente `Ref`** — sigue siendo texto y debe pasar a referencia en la Fase B.")
w("- **Retirada** — marcada para eliminar, todavía presente a propósito.")
w("- **Renombrada** — su nombre actual viene de otro anterior; se indica cuál.")
w("- **Fuera del modelo** — está en la hoja y el modelo objetivo no la contempla.")
w("")
for h in hojas:
    w("### `%s`" % h)
    w("")
    if h in MODELO:
        w(MODELO[h]["proposito"])
    elif h in RETIRADAS:
        w("**Se retira del modelo.** %s" % RETIRADAS[h])
    else:
        w("*No figura en el modelo objetivo.*")
    w("")
    w("%d columnas · %d filas · clave: %s" % (len(encabezados(h)), n_filas[h], muestra(h)))
    w("")
    cols_modelo = {c["nombre"]: c for c in MODELO[h]["columnas"]} if h in MODELO else {}
    w("| # | Columna | Tipo objetivo | Estado |")
    w("|---|---|---|---|")
    for i, c in enumerate(encabezados(h), 1):
        m = cols_modelo.get(c)
        tipo = m["tipo"] if m else "—"
        if m and m.get("pk"):
            tipo += " · **PK**"
        if m and m.get("ref"):
            tipo += " → `%s`" % m["ref"]
        notas = []
        if (h, c) in retipar:
            actual, destino, dest_tabla, _ = retipar[(h, c)]
            notas.append("**Pendiente `Ref`** hacia `%s` (hoy `%s`)" % (dest_tabla, actual))
        if (h, c) in retirados:
            notas.append("**Retirada.** %s" % retirados[(h, c)])
        if (h, c) in origen:
            notas.append(origen[(h, c)])
        if not m and (h, c) not in retirados:
            notas.append("**Fuera del modelo**")
        w("| %d | `%s` | %s | %s |" % (i, c, tipo, " ".join(notas) or ""))
    w("")
    faltan = [n for n in cols_modelo if n not in encabezados(h)]
    if faltan:
        w("**Faltan en la hoja, las declara el modelo:** %s" % ", ".join("`%s`" % x for x in faltan))
        w("")

w("---")
w("*Documento generado. Para actualizarlo, descarga la hoja a `BD/` y ejecuta*")
w("*`python scripts/generar_diccionario_bd.py \"BD/Modelo de Datos (N).xlsx\"`.*")

with open(SALIDA, "w", encoding="utf-8") as f:
    f.write("\n".join(L) + "\n")

print("Generado:", SALIDA)
print("%d hojas, %d filas, %d columnas pendientes de retipar"
      % (len(hojas), sum(n_filas.values()), len(pendientes)))
